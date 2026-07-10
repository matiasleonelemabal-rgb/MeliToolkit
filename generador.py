# -*- coding: utf-8 -*-
"""
Motor de composición de descripciones para publicaciones de MercadoLibre.

Un KIT = lista de (componente_id, cantidad).
Una VARIANTE = ese kit publicado en una CATEGORÍA concreta.

generar_descripcion(kit, categoria) arma el texto completo ordenando el
componente principal primero y usando plantillas fijas para las secciones.
exportar_excel(...) agrupa variantes en hojas, replicando tu flujo manual.
"""

from catalogo import COMPONENTES, CATEGORIAS, nombre_corto, titulo_bloque


# --- Secciones comunes (idénticas en todas las publicaciones) ---------------

BLOQUE_BALIZA_1 = (
    "IMPORTANTE SOBRE LA BALIZA INCLUIDA\n\n"
    "El kit publicado incluye 1 sola baliza reflectiva triangular. Las normativas de "
    "circulación vigentes en Argentina exigen llevar dos (2) balizas portátiles reflectivas "
    "en el automóvil, por lo que esta opción económica está pensada para compradores que ya "
    "cuentan con una baliza adicional en su vehículo."
)

BLOQUE_BALIZA_2 = (
    "IMPORTANTE SOBRE LAS BALIZAS INCLUIDAS\n\n"
    "El kit publicado incluye 2 balizas reflectivas triangulares. Esta versión está pensada "
    "para quienes buscan un kit más completo, con la cantidad de balizas portátiles "
    "reflectivas requeridas para el automóvil.\n\n"
    "Esta aclaración se refiere únicamente a la cantidad de balizas incluidas en el kit."
)

SECCIONES_FINALES = (
    "FACTURACIÓN AUTOMATIZADA\n\n"
    "Emitimos Factura A y B. El proceso está automatizado y tomará directamente los datos "
    "fiscales cargados en tu cuenta al momento de ejecutar la compra.\n\n"
    "LOGÍSTICA Y DESPACHO INMEDIATO\n\n"
    "Envíos Flex: compras previas a las 13:00 hs se entregan en el día, CABA y GBA, entre las "
    "16:00 hs y 22:00 hs.\n\n"
    "Mercado Envíos: despachamos a todo el país en horario comercial con embalaje de alta "
    "densidad para proteger los productos durante el tránsito.\n\n"
    "GARANTÍA\n\n"
    "Nuestros productos cuentan con garantía directa del fabricante por 30 días desde la fecha "
    "de compra."
)


def _lista_ordenada(kit, categoria):
    """Devuelve el kit reordenado con el componente principal primero."""
    principal = CATEGORIAS[categoria]["principal"]
    if principal is None:
        return list(kit)
    ordenado = [x for x in kit if x[0] == principal]
    ordenado += [x for x in kit if x[0] != principal]
    return ordenado


def _titulo(kit, categoria):
    principal = CATEGORIAS[categoria]["principal"]
    tiene_cono = any(cid == "cono" for cid, _ in kit)
    balizas = next((c for cid, c in kit if cid == "baliza"), 0)

    if principal is None:
        extra = []
        if balizas:
            extra.append(f"{balizas} BALIZA" + ("S" if balizas > 1 else ""))
        if tiene_cono:
            extra.append("CONO")
        suf = (" CON " + ", ".join(extra) + " Y AUXILIO VIAL") if extra else " Y AUXILIO VIAL"
        return "KIT DE SEGURIDAD PARA AUTO" + suf

    nombre = nombre_corto(principal, dict(kit)[principal]).upper()
    return f"{nombre} CON KIT DE AUXILIO VIAL"


def _intro(kit, categoria):
    """Plantilla de intro fija según la categoría (ancla semántica)."""
    principal = CATEGORIAS[categoria]["principal"]
    otros = [nombre_corto(cid, c) for cid, c in _lista_ordenada(kit, categoria)
             if cid != principal]
    lista_otros = ", ".join(otros)

    if principal is None:
        items = [nombre_corto(cid, c) for cid, c in kit]
        return (
            "Kit de seguridad para auto y auxilio vial, ideal para llevar en el baúl ante "
            "emergencias, cambio de rueda, recambio de neumáticos o detenciones en ruta, "
            "ciudad, banquina o vía pública. Incluye " + ", ".join(items) + ".\n\n"
            "Este kit reúne herramientas y elementos útiles para responder ante imprevistos "
            "mecánicos, elevar el vehículo, aflojar o ajustar bulones de rueda, proteger las "
            "manos durante la maniobra, mejorar la visibilidad del usuario y señalizar "
            "correctamente una detención de emergencia."
        )

    nombre = nombre_corto(principal, dict(kit)[principal])
    return (
        f"{nombre.capitalize()} con kit de auxilio vial para auto. Esta publicación está "
        f"orientada principalmente al {nombre} como componente central del kit. Además, "
        f"incluye {lista_otros}, formando un conjunto práctico para llevar en el baúl ante "
        "emergencias, cambio de rueda o mantenimiento básico del vehículo."
    )


def generar_descripcion(kit, categoria):
    """kit = [(comp_id, cantidad), ...]  ->  texto completo de la publicación."""
    if categoria not in CATEGORIAS:
        raise ValueError(f"Categoría desconocida: {categoria}")

    orden = _lista_ordenada(kit, categoria)
    partes = []

    # 1) Título
    partes.append(_titulo(kit, categoria))
    partes.append("")

    # 2) Intro
    partes.append(_intro(kit, categoria))
    partes.append("")

    # 3) Aviso de baliza (solo si hay balizas)
    balizas = dict(kit).get("baliza", 0)
    if balizas == 1:
        partes.append(BLOQUE_BALIZA_1)
        partes.append("")
    elif balizas >= 2:
        partes.append(BLOQUE_BALIZA_2)
        partes.append("")

    # 4) Lista "EL KIT PUBLICADO INCLUYE" (principal primero)
    partes.append("EL KIT PUBLICADO INCLUYE\n")
    for cid, c in orden:
        partes.append(COMPONENTES[cid]["inclusion"](c))
    partes.append("")

    # 5) Bloques descriptivos (principal primero)
    for cid, c in orden:
        partes.append(titulo_bloque(cid, c))
        partes.append("")
        partes.append(COMPONENTES[cid]["bloque"](c))
        partes.append("")

    # 6) Usos recomendados (dinámico según componentes)
    usos = "; ".join(COMPONENTES[cid]["uso"] for cid, _ in orden)
    partes.append("USOS RECOMENDADOS\n")
    partes.append(
        f"Este kit es ideal para {usos}, uso preventivo en baúl, seguridad en ruta, "
        "seguridad en ciudad, talleres, cocheras, viajes, vehículos particulares, SUVs "
        "compactos y vehículos ligeros."
    )
    partes.append("")

    # 7) Secciones finales fijas
    partes.append(SECCIONES_FINALES)

    return "\n".join(partes)


# ============================================================
# EXPORT A EXCEL
# ============================================================

def exportar_excel(hojas, ruta_salida):
    """
    hojas = lista de dicts:
        { "nombre": "06_CONO", "kit": [(...)], "categorias": ["kit_seguridad", ...] }
    Cada hoja: una fila por variante (ID, Categoría, Composición, Título, Descripción).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F1F1F")
    header_font = Font(name="Arial", bold=True, color="FFD400")
    cell_font = Font(name="Arial", size=10)
    wrap_top = Alignment(wrap_text=True, vertical="top")

    for hoja in hojas:
        ws = wb.create_sheet(title=hoja["nombre"][:31])
        encabezados = ["ID variante", "Categoría", "Composición", "Título", "Descripción"]
        ws.append(encabezados)
        for col, _ in enumerate(encabezados, start=1):
            celda = ws.cell(row=1, column=col)
            celda.fill = header_fill
            celda.font = header_font

        kit = hoja["kit"]
        composicion = " + ".join(f"{c}x {nombre_corto(cid, c)}" for cid, c in kit)

        for i, cat in enumerate(hoja["categorias"], start=1):
            desc = generar_descripcion(kit, cat)
            titulo = desc.split("\n", 1)[0]
            fila = [
                f"{hoja['nombre']}-{i:02d}",
                CATEGORIAS[cat]["etiqueta"],
                composicion,
                titulo,
                desc,
            ]
            ws.append(fila)

        # formato de columnas
        anchos = [14, 22, 40, 55, 90]
        for col, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[chr(64 + col)].width = ancho
        for row in ws.iter_rows(min_row=2):
            for celda in row:
                celda.font = cell_font
                celda.alignment = wrap_top

    wb.save(ruta_salida)
    return ruta_salida
