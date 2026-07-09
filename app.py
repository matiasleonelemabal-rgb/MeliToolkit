import json
import os
import re
import unicodedata
import zipfile
from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl import load_workbook


# ============================================================
# CONFIGURACIÓN APP
# ============================================================

st.set_page_config(
    page_title="ML Toolkit",
    page_icon="📦",
    layout="wide"
)


# ============================================================
# FUNCIONES COMUNES
# ============================================================

def normalizar_texto(valor):
    if valor is None:
        return ""

    texto = str(valor).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"\s+", " ", texto)

    return texto.lower()


def es_vacio(valor):
    try:
        if pd.isna(valor):
            return True
    except Exception:
        pass

    return valor is None or str(valor).strip() == ""


def resolver_columna_por_nombre(columnas, nombre_buscado):
    buscado = normalizar_texto(nombre_buscado)

    for col in columnas:
        if normalizar_texto(col) == buscado:
            return col

    return None


def detectar_columna(columnas, posibles, obligatorio=True, descripcion="columna"):
    columnas_lista = list(columnas)

    for posible in posibles:
        real = resolver_columna_por_nombre(columnas_lista, posible)
        if real is not None:
            return real

    if obligatorio:
        raise ValueError(
            f"No se encontró {descripcion}. Columnas disponibles: "
            + ", ".join(map(str, columnas_lista))
        )

    return None


def convertir_precio_a_numero(valor):
    """
    Convierte precios tipo:
    - 67285
    - 67.285,00 ARS
    - $ 67.285,00
    - 67285.00
    a número.
    """

    if pd.isna(valor):
        return None

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if texto == "":
        return None

    texto = re.sub(r"[^\d,.\-]", "", texto)

    if texto == "":
        return None

    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "")
            texto = texto.replace(",", ".")
        else:
            texto = texto.replace(",", "")

    elif "," in texto and "." not in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except Exception:
        return None


def convertir_stock(valor):
    numero = convertir_precio_a_numero(valor)

    if numero is None:
        return None

    try:
        return int(float(numero))
    except Exception:
        return None


def formatear_numero(valor):
    if valor is None:
        return None

    try:
        valor = float(valor)
    except Exception:
        return None

    if valor.is_integer():
        return int(valor)

    return round(valor, 2)


def normalizar_sku(valor):
    if pd.isna(valor):
        return ""

    texto = str(valor).strip().upper()

    if texto.endswith(".0"):
        texto = texto[:-2]

    return texto


def obtener_headers_ws(ws):
    headers = {}

    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column

    return headers


def obtener_o_crear_columna(ws, headers, nombre_columna):
    if nombre_columna in headers:
        return headers[nombre_columna]

    nueva_col = ws.max_column + 1
    ws.cell(row=1, column=nueva_col).value = nombre_columna
    headers[nombre_columna] = nueva_col

    return nueva_col


def buscar_hoja(wb, nombre_buscado):
    buscado = normalizar_texto(nombre_buscado)

    for hoja in wb.sheetnames:
        if normalizar_texto(hoja) == buscado:
            return hoja

    return None


def agregar_tabla(ws, titulo, encabezados, registros, fila_vacia):
    ws.append([])
    ws.append([titulo])
    ws.append(encabezados)

    if registros:
        for registro in registros:
            ws.append([registro.get(col, "") for col in encabezados])
    else:
        ws.append(fila_vacia)


def ajustar_anchos(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def reparar_tablas_duplicadas_excel(xlsx_bytes):
    """
    Repara archivos XLSX que tienen tablas internas con nombres duplicados.

    Error típico:
    ValueError: Table with name TablaPublicaciones already exists

    No modifica datos visibles del Excel.
    Solo renombra tablas internas duplicadas dentro del archivo .xlsx.
    """

    entrada = BytesIO(xlsx_bytes)
    salida = BytesIO()

    nombres_usados = set()

    with zipfile.ZipFile(entrada, "r") as zin:
        with zipfile.ZipFile(salida, "w", zipfile.ZIP_DEFLATED) as zout:

            for item in zin.infolist():
                data = zin.read(item.filename)

                if (
                    item.filename.startswith("xl/tables/table")
                    and item.filename.endswith(".xml")
                ):
                    try:
                        texto = data.decode("utf-8")
                    except Exception:
                        zout.writestr(item, data)
                        continue

                    match_display = re.search(r'displayName="([^"]+)"', texto)

                    if match_display:
                        nombre_original = match_display.group(1)
                        nombre_final = nombre_original

                        if nombre_final in nombres_usados:
                            contador = 2

                            while f"{nombre_original}_{contador}" in nombres_usados:
                                contador += 1

                            nombre_final = f"{nombre_original}_{contador}"

                            texto = re.sub(
                                r'displayName="[^"]+"',
                                f'displayName="{nombre_final}"',
                                texto,
                                count=1
                            )

                            texto = re.sub(
                                r'\bname="[^"]+"',
                                f'name="{nombre_final}"',
                                texto,
                                count=1
                            )

                        nombres_usados.add(nombre_final)
                        data = texto.encode("utf-8")

                zout.writestr(item, data)

    salida.seek(0)
    return salida.getvalue()


# ============================================================
# HERRAMIENTA 1 — ACTUALIZAR INTEGRALY
# ============================================================

NOMBRE_HOJA_CONTROL_INTEGRALY = "CONTROL_CRUCE_INTEGRALY"

SUMA_FIJA_PRECIO = 12_000
UMBRAL_PRECIO_INTEGRALY = 1_500_000

ESTADO_ACTIVA = "Activa"
ESTADO_PAUSADA = "Pausada"

POSIBLES_SKU = [
    "sku", "SKU", "Sku",
    "item_code", "ITEM_CODE", "Item Code",
    "codigo", "Código", "CODIGO"
]

POSIBLES_STOCK_ORIGEN = [
    "AStk", "astk", "ASTK",
    "stock", "Stock", "STOCK"
]

POSIBLES_STOCK_DESTINO = [
    "cantidad", "Cantidad", "CANTIDAD",
    "stock", "Stock", "STOCK"
]

POSIBLES_PRECIO_DESTINO = [
    "precio", "Precio", "PRECIO",
    "price", "Price", "PRICE"
]

POSIBLES_ESTADO_DESTINO = [
    "estado", "Estado", "ESTADO",
    "status", "Status", "STATUS"
]

POSIBLES_CUOTAS = [
    "cuotas", "Cuotas", "CUOTAS"
]

POSIBLES_MLA = [
    "mla", "MLA", "Mla",
    "id", "ID",
    "publicacion", "Publicacion", "PUBLICACION"
]

COLUMNAS_PRECIO_ORIGEN = {
    "sin_cuotas": "Precio ML Clasica",
    "3_cuotas": "Precio ML Premium",
    "6_cuotas": "Precio ML Premium 6c",
    "9_cuotas": "Precio ML Premium 9c",
    "12_cuotas": "Precio ML Premium 12c",
}


def clasificar_cuotas(valor):
    if pd.isna(valor):
        return "sin_cuotas"

    texto = str(valor).strip().lower()

    if texto == "":
        return "sin_cuotas"

    texto = (
        texto
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )

    if "no agregar" in texto:
        return "sin_cuotas"

    if "sin cuota" in texto:
        return "sin_cuotas"

    if re.search(r"\b12\b", texto):
        return "12_cuotas"

    if re.search(r"\b9\b", texto):
        return "9_cuotas"

    if re.search(r"\b6\b", texto):
        return "6_cuotas"

    if re.search(r"\b3\b", texto):
        return "3_cuotas"

    return None


def determinar_estado_por_stock(stock_final):
    if stock_final is None:
        return None, "No se pudo determinar stock"

    try:
        stock_num = int(float(stock_final))
    except Exception:
        return None, "Stock inválido"

    if stock_num >= 1:
        return ESTADO_ACTIVA, "Precio actual Integraly menor o igual a 1.500.000 y stock >= 1"

    return ESTADO_PAUSADA, "Precio actual Integraly menor o igual a 1.500.000 y stock <= 0"


def armar_base_global_precios(actualizacion_bytes):
    actualizacion_bytes = reparar_tablas_duplicadas_excel(actualizacion_bytes)

    xls = pd.ExcelFile(BytesIO(actualizacion_bytes))

    filas_origen = []
    control_hojas_ignoradas = []
    control_precios_invalidos = []
    control_stock_invalido = []

    for hoja in xls.sheet_names:

        if str(hoja).upper().startswith("CONTROL"):
            continue

        df = pd.read_excel(
            BytesIO(actualizacion_bytes),
            sheet_name=hoja,
            dtype=str
        )

        if df.empty:
            control_hojas_ignoradas.append({
                "Hoja": hoja,
                "Motivo": "Hoja vacía"
            })
            continue

        columnas = list(df.columns)

        try:
            col_sku = detectar_columna(
                columnas,
                POSIBLES_SKU,
                obligatorio=True,
                descripcion="columna SKU origen"
            )

            col_stock = detectar_columna(
                columnas,
                POSIBLES_STOCK_ORIGEN,
                obligatorio=True,
                descripcion="columna stock origen"
            )

        except Exception as e:
            control_hojas_ignoradas.append({
                "Hoja": hoja,
                "Motivo": str(e)
            })
            continue

        columnas_precio_reales = {}
        faltan_precios = []

        for clave, col_esperada in COLUMNAS_PRECIO_ORIGEN.items():
            col_real = resolver_columna_por_nombre(columnas, col_esperada)

            if col_real is None:
                faltan_precios.append(col_esperada)
            else:
                columnas_precio_reales[clave] = col_real

        if faltan_precios:
            control_hojas_ignoradas.append({
                "Hoja": hoja,
                "Motivo": "Faltan columnas de precio: " + ", ".join(faltan_precios)
            })
            continue

        df["_SKU_KEY"] = df[col_sku].apply(normalizar_sku)
        df = df[df["_SKU_KEY"] != ""].copy()

        for idx, fila in df.iterrows():

            sku_key = fila["_SKU_KEY"]
            stock_final = convertir_stock(fila[col_stock])

            if stock_final is None:
                control_stock_invalido.append({
                    "Hoja origen": hoja,
                    "Fila origen": idx + 2,
                    "SKU": sku_key,
                    "Valor AStk": fila[col_stock],
                })

            registro = {
                "_SKU_KEY": sku_key,
                "_HOJA_ORIGEN": hoja,
                "_FILA_ORIGEN": idx + 2,
                "_ASTK": stock_final,
            }

            for clave_precio, col_precio in columnas_precio_reales.items():

                precio_base = convertir_precio_a_numero(fila[col_precio])

                if precio_base is not None:
                    precio_final = formatear_numero(precio_base + SUMA_FIJA_PRECIO)
                else:
                    precio_final = None

                registro[clave_precio] = precio_final

                if precio_final is None:
                    control_precios_invalidos.append({
                        "Hoja origen": hoja,
                        "Fila origen": idx + 2,
                        "SKU": sku_key,
                        "Columna precio": col_precio,
                        "Valor original": fila[col_precio],
                    })

            filas_origen.append(registro)

    if not filas_origen:
        raise ValueError(
            "No se pudo armar la base de actualización. "
            "Revisá que el archivo tenga SKU, AStk y columnas de precio sin _2."
        )

    df_origen = pd.DataFrame(filas_origen)

    duplicados = df_origen[df_origen["_SKU_KEY"].duplicated(keep=False)].copy()

    df_unico = df_origen.drop_duplicates(
        subset="_SKU_KEY",
        keep="first"
    ).copy()

    datos_por_sku = df_unico.set_index("_SKU_KEY").to_dict(orient="index")

    return (
        datos_por_sku,
        duplicados,
        control_hojas_ignoradas,
        control_precios_invalidos,
        control_stock_invalido
    )


def procesar_integraly(integraly_bytes, actualizacion_bytes):
    (
        datos_por_sku,
        duplicados,
        hojas_ignoradas,
        precios_invalidos,
        stock_invalido
    ) = armar_base_global_precios(actualizacion_bytes)

    integraly_bytes = reparar_tablas_duplicadas_excel(integraly_bytes)
    wb = load_workbook(BytesIO(integraly_bytes))

    if NOMBRE_HOJA_CONTROL_INTEGRALY in wb.sheetnames:
        del wb[NOMBRE_HOJA_CONTROL_INTEGRALY]

    control_resumen = []
    control_sku_sin_match = []
    control_cuotas_no_reconocidas = []
    control_precio_no_actualizado = []
    control_stock_no_actualizado = []
    control_estado_actualizado = []
    control_estado_no_actualizado = []
    control_mayor_umbral_pausada = []
    control_bloqueados_por_umbral_integraly = []

    for ws in wb.worksheets:

        if ws.title.upper().startswith("CONTROL"):
            continue

        headers = obtener_headers_ws(ws)

        col_sku_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_SKU,
            obligatorio=False,
            descripcion="columna SKU destino"
        )

        if col_sku_nombre is None:
            control_resumen.append({
                "Hoja destino": ws.title,
                "Estado hoja": "Ignorada",
                "Motivo": "No se encontró columna SKU",
                "Filas procesadas": 0,
                "Filas actualizadas con precio": 0,
                "Filas actualizadas con stock": 0,
                "SKU sin match": 0,
                "Cuotas no reconocidas": 0,
                "Precio no actualizado": 0,
                "Stock no actualizado": 0,
                "Estado actualizado": 0,
                "Estado no actualizado": 0,
                "Pausadas por precio Integraly mayor a 1.500.000": 0,
                "Bloqueadas sin modificar precio/stock": 0,
            })
            continue

        col_mla_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_MLA,
            obligatorio=False,
            descripcion="columna MLA"
        )

        col_cuotas_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_CUOTAS,
            obligatorio=False,
            descripcion="columna cuotas"
        )

        col_precio_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_PRECIO_DESTINO,
            obligatorio=False,
            descripcion="columna precio destino"
        )

        col_stock_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_STOCK_DESTINO,
            obligatorio=False,
            descripcion="columna stock destino"
        )

        col_estado_nombre = detectar_columna(
            headers.keys(),
            POSIBLES_ESTADO_DESTINO,
            obligatorio=False,
            descripcion="columna estado destino"
        )

        if col_precio_nombre is None:
            col_precio_nombre = "precio"
            obtener_o_crear_columna(ws, headers, col_precio_nombre)

        if col_stock_nombre is None:
            col_stock_nombre = "cantidad"
            obtener_o_crear_columna(ws, headers, col_stock_nombre)

        if col_estado_nombre is None:
            col_estado_nombre = "estado"
            obtener_o_crear_columna(ws, headers, col_estado_nombre)

        col_sku = headers[col_sku_nombre]
        col_precio = headers[col_precio_nombre]
        col_stock = headers[col_stock_nombre]
        col_estado = headers[col_estado_nombre]

        col_mla = headers[col_mla_nombre] if col_mla_nombre else None
        col_cuotas = headers[col_cuotas_nombre] if col_cuotas_nombre else None

        filas_procesadas = 0
        filas_actualizadas_precio = 0
        filas_actualizadas_stock = 0
        sku_sin_match = 0
        cuotas_no_reconocidas = 0
        precio_no_actualizado = 0
        stock_no_actualizado = 0
        estado_actualizado = 0
        estado_no_actualizado = 0
        pausadas_umbral_integraly = 0
        bloqueadas_umbral_integraly = 0

        for row in range(2, ws.max_row + 1):

            sku_original = ws.cell(row=row, column=col_sku).value
            sku_key = normalizar_sku(sku_original)

            if sku_key == "":
                continue

            filas_procesadas += 1

            mla = ws.cell(row=row, column=col_mla).value if col_mla else ""
            cuotas_original = ws.cell(row=row, column=col_cuotas).value if col_cuotas else ""
            estado_anterior = ws.cell(row=row, column=col_estado).value
            precio_anterior = ws.cell(row=row, column=col_precio).value
            stock_anterior = ws.cell(row=row, column=col_stock).value

            precio_integraly_actual_num = convertir_precio_a_numero(precio_anterior)

            if sku_key not in datos_por_sku:
                sku_sin_match += 1
                control_sku_sin_match.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Cuotas": cuotas_original,
                    "Precio actual Integraly": precio_anterior,
                    "Estado anterior": estado_anterior,
                })
                continue

            datos = datos_por_sku[sku_key]
            stock_final = datos.get("_ASTK")

            clave_precio = clasificar_cuotas(cuotas_original)

            if clave_precio is None:
                cuotas_no_reconocidas += 1
                estado_no_actualizado += 1

                control_cuotas_no_reconocidas.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Cuotas": cuotas_original,
                    "Precio actual Integraly": precio_anterior,
                })

                control_estado_no_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": "",
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Motivo": "Cuotas no reconocidas, no se pudo determinar precio",
                })

                continue

            precio_final = datos.get(clave_precio)

            if precio_final is None:
                precio_no_actualizado += 1
                estado_no_actualizado += 1

                control_precio_no_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Cuotas": cuotas_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio requerido": COLUMNAS_PRECIO_ORIGEN[clave_precio],
                    "Hoja origen": datos.get("_HOJA_ORIGEN"),
                    "Fila origen": datos.get("_FILA_ORIGEN"),
                })

                control_estado_no_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": precio_final,
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Motivo": "Precio final Global + 12000 inválido o vacío",
                })

                continue

            # ====================================================
            # REGLA CRÍTICA:
            # SI EL PRECIO ACTUAL DE INTEGRALY ES MAYOR A 1.500.000:
            # - NO MODIFICA PRECIO
            # - NO MODIFICA STOCK
            # - SOLO MODIFICA ESTADO A PAUSADA
            # ====================================================

            if (
                precio_integraly_actual_num is not None
                and precio_integraly_actual_num > UMBRAL_PRECIO_INTEGRALY
            ):

                ws.cell(row=row, column=col_estado).value = ESTADO_PAUSADA

                estado_actualizado += 1
                pausadas_umbral_integraly += 1
                bloqueadas_umbral_integraly += 1

                registro_bloqueado = {
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Cuotas": cuotas_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio actual Integraly numérico": precio_integraly_actual_num,
                    "Stock actual Integraly": stock_anterior,
                    "Precio calculado Global + 12000": precio_final,
                    "Stock Global AStk": stock_final,
                    "Estado anterior": estado_anterior,
                    "Estado final": ESTADO_PAUSADA,
                    "Motivo": "Precio actual de Integraly mayor a 1.500.000: no se modifica precio ni stock, solo se pausa",
                }

                control_bloqueados_por_umbral_integraly.append(registro_bloqueado)

                control_mayor_umbral_pausada.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": precio_final,
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Estado final": ESTADO_PAUSADA,
                    "Motivo": "Precio actual de Integraly mayor a 1.500.000: no se modifica precio ni stock, solo se pausa",
                })

                control_estado_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": precio_final,
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Estado final": ESTADO_PAUSADA,
                    "Motivo": "Precio actual de Integraly mayor a 1.500.000: no se modifica precio ni stock, solo se pausa",
                })

                continue

            ws.cell(row=row, column=col_precio).value = precio_final
            filas_actualizadas_precio += 1

            if stock_final is not None:
                ws.cell(row=row, column=col_stock).value = stock_final
                filas_actualizadas_stock += 1
            else:
                stock_no_actualizado += 1
                control_stock_no_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Valor AStk": stock_final,
                    "Hoja origen": datos.get("_HOJA_ORIGEN"),
                    "Fila origen": datos.get("_FILA_ORIGEN"),
                })

            estado_final, motivo_estado = determinar_estado_por_stock(stock_final)

            if estado_final is not None:
                ws.cell(row=row, column=col_estado).value = estado_final
                estado_actualizado += 1

                control_estado_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": precio_final,
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Estado final": estado_final,
                    "Motivo": motivo_estado,
                })
            else:
                estado_no_actualizado += 1

                control_estado_no_actualizado.append({
                    "Hoja destino": ws.title,
                    "Fila destino": row,
                    "MLA": mla,
                    "SKU": sku_original,
                    "Precio actual Integraly": precio_anterior,
                    "Precio final Global + 12000": precio_final,
                    "Stock final": stock_final,
                    "Estado anterior": estado_anterior,
                    "Motivo": motivo_estado,
                })

        control_resumen.append({
            "Hoja destino": ws.title,
            "Estado hoja": "Procesada",
            "Motivo": "",
            "Filas procesadas": filas_procesadas,
            "Filas actualizadas con precio": filas_actualizadas_precio,
            "Filas actualizadas con stock": filas_actualizadas_stock,
            "SKU sin match": sku_sin_match,
            "Cuotas no reconocidas": cuotas_no_reconocidas,
            "Precio no actualizado": precio_no_actualizado,
            "Stock no actualizado": stock_no_actualizado,
            "Estado actualizado": estado_actualizado,
            "Estado no actualizado": estado_no_actualizado,
            "Pausadas por precio Integraly mayor a 1.500.000": pausadas_umbral_integraly,
            "Bloqueadas sin modificar precio/stock": bloqueadas_umbral_integraly,
        })

    ws_control = wb.create_sheet(NOMBRE_HOJA_CONTROL_INTEGRALY)

    ws_control.append(["Fecha proceso", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_control.append(["Suma fija aplicada a cada precio Global", SUMA_FIJA_PRECIO])
    ws_control.append(["Umbral precio actual Integraly", UMBRAL_PRECIO_INTEGRALY])
    ws_control.append(["Estado activa", ESTADO_ACTIVA])
    ws_control.append(["Estado pausada", ESTADO_PAUSADA])
    ws_control.append([])

    ws_control.append(["MAPEO DE CUOTAS"])
    ws_control.append(["Condición en Integraly", "Columna usada en Global"])
    ws_control.append(["Sin cuotas / No Agregar Cuotas / vacío", "Precio ML Clasica + 12000"])
    ws_control.append(["3 cuotas", "Precio ML Premium + 12000"])
    ws_control.append(["6 cuotas", "Precio ML Premium 6c + 12000"])
    ws_control.append(["9 cuotas", "Precio ML Premium 9c + 12000"])
    ws_control.append(["12 cuotas", "Precio ML Premium 12c + 12000"])
    ws_control.append(["Stock", "AStk"])
    ws_control.append([])

    ws_control.append(["REGLA CRÍTICA PRECIO ACTUAL INTEGRALY"])
    ws_control.append(["Condición", "Acción"])
    ws_control.append(["Precio actual Integraly > 1.500.000", "NO modifica precio, NO modifica stock, SOLO estado = Pausada"])
    ws_control.append(["Precio actual Integraly <= 1.500.000 y stock Global >= 1", "Modifica precio, modifica stock, estado = Activa"])
    ws_control.append(["Precio actual Integraly <= 1.500.000 y stock Global <= 0", "Modifica precio, modifica stock, estado = Pausada"])
    ws_control.append(["Precio actual Integraly vacío/no legible", "Modifica precio y stock si hay match; estado según stock"])
    ws_control.append([])

    ws_control.append(["RESUMEN"])

    resumen_cols = [
        "Hoja destino",
        "Estado hoja",
        "Motivo",
        "Filas procesadas",
        "Filas actualizadas con precio",
        "Filas actualizadas con stock",
        "SKU sin match",
        "Cuotas no reconocidas",
        "Precio no actualizado",
        "Stock no actualizado",
        "Estado actualizado",
        "Estado no actualizado",
        "Pausadas por precio Integraly mayor a 1.500.000",
        "Bloqueadas sin modificar precio/stock",
    ]

    ws_control.append(resumen_cols)

    for r in control_resumen:
        ws_control.append([r.get(c, "") for c in resumen_cols])

    agregar_tabla(
        ws_control,
        "BLOQUEADAS POR PRECIO ACTUAL INTEGRALY MAYOR A 1.500.000 - NO SE MODIFICÓ PRECIO NI STOCK",
        [
            "Hoja destino",
            "Fila destino",
            "MLA",
            "SKU",
            "Cuotas",
            "Precio actual Integraly",
            "Precio actual Integraly numérico",
            "Stock actual Integraly",
            "Precio calculado Global + 12000",
            "Stock Global AStk",
            "Estado anterior",
            "Estado final",
            "Motivo",
        ],
        control_bloqueados_por_umbral_integraly,
        ["Sin bloqueadas por precio Integraly mayor a 1.500.000", "", "", "", "", "", "", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "SKU SIN MATCH",
        ["Hoja destino", "Fila destino", "MLA", "SKU", "Cuotas", "Precio actual Integraly", "Estado anterior"],
        control_sku_sin_match,
        ["Sin SKU sin match", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "CUOTAS NO RECONOCIDAS",
        ["Hoja destino", "Fila destino", "MLA", "SKU", "Cuotas", "Precio actual Integraly"],
        control_cuotas_no_reconocidas,
        ["Sin cuotas no reconocidas", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "PRECIO NO ACTUALIZADO",
        ["Hoja destino", "Fila destino", "MLA", "SKU", "Cuotas", "Precio actual Integraly", "Precio requerido", "Hoja origen", "Fila origen"],
        control_precio_no_actualizado,
        ["Sin precios pendientes", "", "", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "STOCK NO ACTUALIZADO",
        ["Hoja destino", "Fila destino", "MLA", "SKU", "Valor AStk", "Hoja origen", "Fila origen"],
        control_stock_no_actualizado,
        ["Sin stocks pendientes", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "ESTADO ACTUALIZADO",
        [
            "Hoja destino",
            "Fila destino",
            "MLA",
            "SKU",
            "Precio actual Integraly",
            "Precio final Global + 12000",
            "Stock final",
            "Estado anterior",
            "Estado final",
            "Motivo",
        ],
        control_estado_actualizado,
        ["Sin estados actualizados", "", "", "", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "ESTADO NO ACTUALIZADO",
        [
            "Hoja destino",
            "Fila destino",
            "MLA",
            "SKU",
            "Precio actual Integraly",
            "Precio final Global + 12000",
            "Stock final",
            "Estado anterior",
            "Motivo",
        ],
        control_estado_no_actualizado,
        ["Sin estados pendientes", "", "", "", "", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "PAUSADAS POR PRECIO ACTUAL INTEGRALY MAYOR A 1.500.000",
        [
            "Hoja destino",
            "Fila destino",
            "MLA",
            "SKU",
            "Precio actual Integraly",
            "Precio final Global + 12000",
            "Stock final",
            "Estado anterior",
            "Estado final",
            "Motivo",
        ],
        control_mayor_umbral_pausada,
        ["Sin publicaciones mayores a 1.500.000 pausadas", "", "", "", "", "", "", "", "", ""]
    )

    ws_control.append([])
    ws_control.append(["SKU DUPLICADOS EN GLOBAL / ACTUALIZACIÓN PRECIO"])
    ws_control.append(["SKU", "Cantidad apariciones", "Hojas origen"])

    if not duplicados.empty:
        dup_resumen = (
            duplicados
            .groupby("_SKU_KEY")
            .agg(
                cantidad_apariciones=("_SKU_KEY", "size"),
                hojas_origen=("_HOJA_ORIGEN", lambda x: " | ".join(sorted(set(map(str, x)))))
            )
            .reset_index()
        )

        for _, r in dup_resumen.iterrows():
            ws_control.append([
                r["_SKU_KEY"],
                int(r["cantidad_apariciones"]),
                r["hojas_origen"]
            ])
    else:
        ws_control.append(["Sin duplicados", 0, ""])

    agregar_tabla(
        ws_control,
        "HOJAS IGNORADAS EN GLOBAL / ACTUALIZACIÓN PRECIO",
        ["Hoja", "Motivo"],
        hojas_ignoradas,
        ["Sin hojas ignoradas", ""]
    )

    agregar_tabla(
        ws_control,
        "PRECIOS INVÁLIDOS EN GLOBAL / ACTUALIZACIÓN PRECIO",
        ["Hoja origen", "Fila origen", "SKU", "Columna precio", "Valor original"],
        precios_invalidos,
        ["Sin precios inválidos", "", "", "", ""]
    )

    agregar_tabla(
        ws_control,
        "STOCK INVÁLIDO EN GLOBAL / ACTUALIZACIÓN PRECIO",
        ["Hoja origen", "Fila origen", "SKU", "Valor AStk"],
        stock_invalido,
        ["Sin stocks inválidos", "", "", ""]
    )

    ajustar_anchos(ws_control)

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)

    resumen = {
        "control_resumen": control_resumen,
        "sku_unicos_origen": len(datos_por_sku),
        "sku_duplicados_origen": duplicados["_SKU_KEY"].nunique() if not duplicados.empty else 0,
        "hojas_ignoradas_origen": len(hojas_ignoradas),
    }

    return salida, resumen


def mostrar_herramienta_integraly():
    st.title("Actualizar Integraly")
    st.caption("Actualiza precio, stock y estado por SKU.")

    st.warning(
        "La herramienta toma los precios del archivo Global, limpia ARS, convierte a número "
        "y suma $12.000 fijos antes de completar Integraly."
    )

    st.error(
        "Regla crítica: si el precio actual de Integraly supera $1.500.000, "
        "NO se modifica precio ni stock. Solo se fuerza el estado a Pausada."
    )

    st.info(
        "Mapeo: Clásica = sin cuotas | Premium = 3 cuotas | Premium 6c = 6 cuotas | "
        "Premium 9c = 9 cuotas | Premium 12c = 12 cuotas."
    )

    col1, col2 = st.columns(2)

    with col1:
        archivo_integraly = st.file_uploader(
            "Subí el archivo Integraly",
            type=["xlsx"],
            key="integraly_uploader"
        )

    with col2:
        archivo_actualizacion = st.file_uploader(
            "Subí el archivo Global / Actualización Precio",
            type=["xlsx"],
            key="actualizacion_uploader"
        )

    if st.button("Procesar Integraly", type="primary"):

        if archivo_integraly is None or archivo_actualizacion is None:
            st.error("Tenés que subir los 2 archivos para procesar.")

        else:
            try:
                with st.spinner("Procesando archivos..."):
                    salida, resumen = procesar_integraly(
                        archivo_integraly.getvalue(),
                        archivo_actualizacion.getvalue()
                    )

                st.success("Proceso finalizado correctamente.")

                st.subheader("Resumen")
                st.dataframe(
                    pd.DataFrame(resumen["control_resumen"]),
                    use_container_width=True
                )

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("SKU únicos origen", resumen["sku_unicos_origen"])
                c2.metric("SKU duplicados origen", resumen["sku_duplicados_origen"])
                c3.metric("Hojas ignoradas origen", resumen["hojas_ignoradas_origen"])
                c4.metric("Suma fija aplicada", f"${SUMA_FIJA_PRECIO:,.0f}")

                fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
                nombre_salida = f"INTEGRALY_PRECIOS_STOCK_ESTADO_ACTUALIZADO_{fecha}.xlsx"

                st.download_button(
                    label="Descargar Excel Integraly actualizado",
                    data=salida,
                    file_name=nombre_salida,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as e:
                st.error("El proceso falló.")
                st.exception(e)


# ============================================================
# ============================================================
# HERRAMIENTA 2 — AGENTE PUBLICADOR
# ============================================================
# Solo pide el Excel Publicar de Meli.
# Detecta las categorías disponibles por nombre de hoja.
# Los campos fijos y los "solicitar" están embebidos en el código
# (parseados del AGENTE_PUBLICADOR.xlsx original).
# Modifica el archivo in-place (desmergea fila 8 si hace falta)
# para que Meli acepte el mismo archivo de vuelta.
# ============================================================

FILA_INICIO_PUBLICAR = 8
FILA_HEADERS_PUBLICAR = 3

# Fuente de verdad persistente en Streamlit Community Cloud: un archivo
# COMMITEADO en el repo. El disco del contenedor es efímero (se borra al
# dormirse/redeploy), así que la persistencia real la da el repo, no una
# escritura en runtime.
#
# Para agregar categorías: reemplazá este archivo en GitHub y hacé commit.
#
# Prioridad de lectura:
#   1) lo que subiste por la UI en esta sesión (override temporal)
#   2) el archivo committeado en el repo  <-- persistente
#   3) una copia escrita en runtime (solo ayuda dentro de una misma sesión)
_APP_DIR = os.path.dirname(os.path.abspath(__file__))

# Rutas candidatas del AGENTE committeado (se usa la primera que exista)
_AGENTE_REPO_CANDIDATES = [
    os.path.join(_APP_DIR, "AGENTE_PUBLICADOR.xlsx"),
    os.path.join(_APP_DIR, "data", "AGENTE_PUBLICADOR.xlsx"),
    os.path.join(_APP_DIR, "assets", "AGENTE_PUBLICADOR.xlsx"),
]

# Copia en runtime (efímera en la nube; útil solo dentro de la sesión)
_AGENTE_PATH = os.path.join(_APP_DIR, "AGENTE_PUBLICADOR_guardado.xlsx")


def _ruta_agente_repo():
    """Devuelve la ruta del AGENTE committeado si existe, si no None."""
    for ruta in _AGENTE_REPO_CANDIDATES:
        if os.path.exists(ruta):
            return ruta
    return None


def _guardar_agente_ap(archivo_bytes):
    # En memoria: override para esta sesión
    st.session_state["agente_publicador_bytes"] = archivo_bytes
    # Intento de copia en disco (persiste solo si el disco no es efímero)
    try:
        with open(_AGENTE_PATH, "wb") as f:
            f.write(archivo_bytes)
        st.session_state["agente_persistido"] = True
    except OSError:
        st.session_state["agente_persistido"] = False


def _agente_existe():
    if st.session_state.get("agente_publicador_bytes") is not None:
        return True
    if _ruta_agente_repo() is not None:
        return True
    return os.path.exists(_AGENTE_PATH)


def _leer_agente_ap():
    # 1) Override de la sesión
    en_memoria = st.session_state.get("agente_publicador_bytes")
    if en_memoria is not None:
        return en_memoria
    # 2) Archivo committeado en el repo (fuente persistente)
    ruta_repo = _ruta_agente_repo()
    if ruta_repo is not None:
        try:
            with open(ruta_repo, "rb") as f:
                return f.read()
        except OSError:
            pass
    # 3) Copia en runtime (efímera)
    if os.path.exists(_AGENTE_PATH):
        try:
            with open(_AGENTE_PATH, "rb") as f:
                return f.read()
        except OSError:
            return None
    return None


def _origen_agente():
    """Para la UI: de dónde sale el AGENTE que se está usando."""
    if st.session_state.get("agente_publicador_bytes") is not None:
        return "sesion"
    if _ruta_agente_repo() is not None:
        return "repo"
    if os.path.exists(_AGENTE_PATH):
        return "disco"
    return None


@st.cache_data(show_spinner=False)
def _sheetnames_publicar_ap(pub_bytes):
    """Lee solo los nombres de hoja del Publicar (read_only, cacheado)."""
    w = load_workbook(BytesIO(pub_bytes), read_only=True)
    nombres = list(w.sheetnames)
    w.close()
    return nombres

# Mapeo: nombre de categoría en el AGENTE -> nombre de hoja en Publicar de Meli
MAPEO_CATEGORIAS_AP = {
    "Kit de Seguridad para Autos": "Kit de Seguridad para Autos",
    "Crique Hidráulico": "Hidráulicos",
    "Llave Cruz / Llave de Rueda": "Llaves de Cruz",
    "Chaleco Reflectivo": "Chalecos Reflectivos",
    "Baliza": "Balizas",
    "Cono de Seguridad Vial": "Conos de Seguridad",
    "Compresor de Aire": "Compresores",
}
# Mapa con claves normalizadas (insensible a mayúsculas/acentos)
MAPEO_CATEGORIAS_AP_NORM = {normalizar_texto(k): v for k, v in MAPEO_CATEGORIAS_AP.items()}

_STOPWORDS_AP = {"de", "la", "el", "para", "y", "a"}


def _tokens_categoria_ap(texto):
    """Tokens normalizados, sin stopwords y con la 's' final del plural quitada."""
    tokens = set()
    for palabra in normalizar_texto(texto).split():
        if palabra in _STOPWORDS_AP:
            continue
        if len(palabra) > 3 and palabra.endswith("s"):
            palabra = palabra[:-1]
        tokens.add(palabra)
    return tokens


# Overrides por categoría: columnas que se FUERZAN a "solicitar" (pedir desde
# la app) aunque en el AGENTE tengan un valor fijo. Clave: categoría normalizada
# -> { header normalizado: tipo }. tipo: "texto" (input) o "select_sino"
# (desplegable Sí/No). El match del header es por IGUALDAD EXACTA normalizada
# (no substring), para no tocar columnas como el Título o "Modelo detallado".
OVERRIDES_SOLICITAR_AP = {
    "kit de seguridad para auto": {
        "numero de pieza": "texto",
        "balizas reflactarias": "select_sino",
        "cantidad de piezas": "texto",
    },
    "crique hidraulico": {
        "modelo": "texto",
        "fotos": "texto",
    },
    "llave cruz / llave de rueda": {
        "numero de pieza": "texto",
        "modelo": "texto",
    },
    "chaleco reflectivo": {
        "modelo": "texto",
    },
    "baliza": {
        "numero de pieza": "texto",
        "modelo": "texto",
    },
    "cono de seguridad vial": {
        "modelo": "texto",
    },
    "compresor de aire": {
        "modelo": "texto",
    },
}


@st.cache_data(show_spinner=False)
def parsear_agente_publicador_ap(archivo_bytes):
    """
    Lee la hoja 'celdas a completar' del AGENTE_PUBLICADOR.xlsx y
    devuelve los bloques con campos fijos y solicitar, cada uno ya
    mapeado al header name normalizado para cruzar con el Publicar.
    """
    wb = load_workbook(BytesIO(archivo_bytes), read_only=True, data_only=True)

    hoja_datos = None
    for h in wb.sheetnames:
        hn = normalizar_texto(h)
        if "completar" in hn or "publicar_base" in hn or "categoria" in hn:
            hoja_datos = h
            break

    # Fallback: si hay una sola hoja, usarla (el parser escanea por contenido)
    if hoja_datos is None and len(wb.sheetnames) == 1:
        hoja_datos = wb.sheetnames[0]

    if hoja_datos is None:
        raise ValueError(
            "El AGENTE no tiene una hoja reconocible "
            "('celdas a completar' / 'publicar_base_x_categoria')."
        )

    ws = wb[hoja_datos]
    max_col = 55
    max_fila = 80

    # Encontrar filas de título de categoría
    titulos = []
    for i in range(1, max_fila):
        v = ws.cell(i, 1).value
        if not v:
            continue
        t = str(v).strip()
        if len(t) > 80:
            continue
        if re.match(r'^CATEGOR[ÍI]A\b', t, re.IGNORECASE) or re.match(r'^COMPRESOR\b', t, re.IGNORECASE):
            # El SKU que diferencia bloques de la misma categoría suele estar
            # al lado del título (ej. col B: "SKU: COM001"). Se escanea la fila.
            sku = None
            for c in range(1, max_col):
                celda = ws.cell(i, c).value
                if celda and isinstance(celda, str):
                    m = re.search(r'sku\s*[:\-]?\s*([A-Za-z0-9._\-]+)', celda, re.IGNORECASE)
                    if m:
                        sku = m.group(1).strip()
                        break
            titulos.append((i, t, sku))

    bloques = []

    for idx, (fila_titulo, nombre_cat, sku_bloque) in enumerate(titulos):
        limite = titulos[idx + 1][0] if idx + 1 < len(titulos) else max_fila

        # Header = primera fila con "Código de catálogo"
        fila_header = None
        for check in range(fila_titulo + 1, limite):
            v = ws.cell(check, 1).value
            if v and 'código' in str(v).lower() and 'catálogo' in str(v).lower():
                fila_header = check
                break
        if not fila_header:
            continue

        # Datos = primera fila con "solicitar"
        fila_datos = None
        for check in range(fila_header + 1, limite):
            vals = [ws.cell(check, c).value for c in range(1, max_col)]
            if any(v and isinstance(v, str) and 'solicitar' in v.lower() for v in vals):
                fila_datos = check
                break
        if not fila_datos:
            continue

        headers = [ws.cell(fila_header, c).value for c in range(1, max_col)]
        datos = [ws.cell(fila_datos, c).value for c in range(1, max_col)]

        campos_solicitar = []
        campos_fijos = []
        mapa_header_texto = {}  # header_norm -> texto original (para labels)

        for c, (h, d) in enumerate(zip(headers, datos), start=1):
            if h is None and d is None:
                continue
            h_str = str(h).strip() if h else ""
            h_norm = normalizar_texto(h_str)
            if h_norm:
                mapa_header_texto[h_norm] = h_str

            if 'buybox' in h_norm or 'hidden' in h_norm:
                continue

            if d is not None and isinstance(d, str) and 'solicitar' in d.lower():
                label = h_str
                es_titulo = ('titulo' in h_norm or 'título' in h_str.lower())
                if es_titulo:
                    label = "Títulos"
                elif len(label) > 40:
                    label = label[:40]

                campos_solicitar.append({
                    "header_norm": h_norm,
                    "label": label,
                    "es_titulo": es_titulo,
                    "tipo": "texto",
                })
            elif d is not None:
                campos_fijos.append({
                    "header_norm": h_norm,
                    "valor": d,
                })

        # Nombre limpio para mostrar
        nombre_limpio = re.sub(r'^CATEGOR[ÍI]A\s+', '', nombre_cat, flags=re.IGNORECASE).strip()

        # --- Overrides: forzar ciertas columnas a "solicitar" ---
        override = OVERRIDES_SOLICITAR_AP.get(normalizar_texto(nombre_limpio), {})
        for hn_target, tipo in override.items():
            # Si ya es un campo solicitar, solo fijo su tipo (ej. select Sí/No)
            ya = next((c for c in campos_solicitar if c["header_norm"] == hn_target), None)
            if ya is not None:
                ya["tipo"] = tipo
                continue
            # Si estaba como fijo, lo saco de fijos y lo paso a solicitar
            fijo = next((c for c in campos_fijos if c["header_norm"] == hn_target), None)
            if fijo is not None:
                campos_fijos.remove(fijo)
            # Construyo el campo solicitar (aunque el dato estuviera vacío)
            if hn_target in mapa_header_texto:
                label = mapa_header_texto[hn_target]
                if len(label) > 40:
                    label = label[:40]
                campos_solicitar.append({
                    "header_norm": hn_target,
                    "label": label,
                    "es_titulo": False,
                    "tipo": tipo,
                })

        bloques.append({
            "nombre_original": nombre_cat,
            "nombre_limpio": nombre_limpio,
            "sku": sku_bloque,
            "campos_solicitar": campos_solicitar,
            "campos_fijos": campos_fijos,
        })

    # Etiquetas únicas. Si hay varios bloques de la misma categoría, se
    # diferencian por su SKU (ej. "Compresor de Aire — COM001"). Solo si
    # falta el SKU se cae al numerito genérico "(variante N)".
    conteo = {}
    for b in bloques:
        conteo[b["nombre_limpio"]] = conteo.get(b["nombre_limpio"], 0) + 1
    indices = {}
    for b in bloques:
        n = b["nombre_limpio"]
        if conteo[n] > 1:
            if b.get("sku"):
                b["etiqueta"] = f"{n} — {b['sku']}"
            else:
                indices[n] = indices.get(n, 0) + 1
                b["etiqueta"] = f"{n} (variante {indices[n]})"
        else:
            # Único en su categoría: igual mostramos el SKU si existe.
            b["etiqueta"] = f"{n} — {b['sku']}" if b.get("sku") else n

    wb.close()
    return bloques


def desmerguear_fila(ws, fila):
    """Quita todos los merges que incluyan esta fila, para poder escribir."""
    merges_a_quitar = [
        m for m in ws.merged_cells.ranges
        if m.min_row <= fila <= m.max_row
    ]
    for m in merges_a_quitar:
        ws.unmerge_cells(str(m))


def encontrar_hoja_publicar(sheetnames, nombre_limpio):
    """
    Devuelve la hoja del Publicar que corresponde a esta categoría.
    Orden: mapa normalizado -> igualdad normalizada -> tokens con plural
    tolerante. Recibe una lista de sheetnames (no el wb) para no forzar
    cargas pesadas del Excel en cada rerun.
    """
    ignorar = {normalizar_texto("Ayuda"), normalizar_texto("extra info")}
    hojas = [h for h in sheetnames if normalizar_texto(h) not in ignorar]

    # 1) Mapa conocido, insensible a mayúsculas/acentos
    esperada = MAPEO_CATEGORIAS_AP_NORM.get(normalizar_texto(nombre_limpio))
    if esperada:
        for h in hojas:
            if normalizar_texto(h) == normalizar_texto(esperada):
                return h

    # 2) Igualdad normalizada directa
    for h in hojas:
        if normalizar_texto(h) == normalizar_texto(nombre_limpio):
            return h

    # 3) Match por tokens, tolerante a singular/plural y orden
    tc = _tokens_categoria_ap(nombre_limpio)
    if tc:
        for h in hojas:
            th = _tokens_categoria_ap(h)
            if th and (tc <= th or th <= tc):
                return h

    return None


def mapear_headers_publicar(ws):
    """
    Lee los headers de fila 3 del Publicar y devuelve
    {header_normalizado: número_de_columna}.
    """
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(FILA_HEADERS_PUBLICAR, c).value
        if h:
            headers[normalizar_texto(h)] = c
    return headers


def escribir_publicaciones_ap(wb, hoja_nombre, bloque, titulos, valores_solicitados):
    """
    Escribe N filas en la hoja del Publicar (modificando in-place).
    Si fila 8 tiene merges, arranca desde fila 9 para no romper la
    estructura que Meli valida al importar. No toca BUYBOX ni HIDDEN.
    """
    ws = wb[hoja_nombre]
    headers_pub = mapear_headers_publicar(ws)

    # Si fila 8 tiene merges, arrancar desde fila 9
    tiene_merges_f8 = any(
        m.min_row <= FILA_INICIO_PUBLICAR <= m.max_row
        for m in ws.merged_cells.ranges
    )
    fila_inicio_real = (FILA_INICIO_PUBLICAR + 1) if tiene_merges_f8 else FILA_INICIO_PUBLICAR

    # Encontrar columna de título y cantidad de caracteres
    col_titulo = None
    col_caracteres = None
    for h_norm, col in headers_pub.items():
        if 'titulo' in h_norm or 'título' in h_norm:
            col_titulo = col
        if 'cantidad de caracteres' in h_norm:
            col_caracteres = col

    # Columnas que NO se tocan
    cols_no_tocar = set()
    for h_norm, col in headers_pub.items():
        if 'buybox' in h_norm or 'hidden' in h_norm:
            cols_no_tocar.add(col)

    for i, titulo in enumerate(titulos):
        fila = fila_inicio_real + i

        # Fijos
        for campo in bloque["campos_fijos"]:
            col = headers_pub.get(campo["header_norm"])
            if col and col not in cols_no_tocar:
                ws.cell(fila, col).value = campo["valor"]

        # Solicitados
        for campo in bloque["campos_solicitar"]:
            if campo["es_titulo"]:
                if col_titulo:
                    ws.cell(fila, col_titulo).value = titulo
                if col_caracteres:
                    ws.cell(fila, col_caracteres).value = len(titulo)
            else:
                col = headers_pub.get(campo["header_norm"])
                if col and col not in cols_no_tocar:
                    valor = valores_solicitados.get(campo["label"], "")
                    ws.cell(fila, col).value = valor

    return len(titulos)


def mostrar_herramienta_agente_publicador():
    st.title("Agente Publicador")
    st.caption(
        "Subí el Excel Publicar de Meli. La app detecta las categorías por "
        "las hojas del archivo, te pide solo los campos variables, y te "
        "devuelve el mismo archivo completado."
    )

    # --- AGENTE: fuente de verdad = archivo committeado en el repo ---
    origen = _origen_agente()
    ya_guardado = origen is not None
    titulo_exp = (
        "AGENTE_PUBLICADOR ya configurado (tocá para reemplazar por esta sesión)"
        if ya_guardado
        else "Configurar AGENTE_PUBLICADOR"
    )
    with st.expander(titulo_exp, expanded=not ya_guardado):
        st.caption(
            "Persistencia en Streamlit Cloud: el AGENTE tiene que estar "
            "committeado en el repo (AGENTE_PUBLICADOR.xlsx en la raíz, o en "
            "data/ o assets/). Para agregar categorías, reemplazá ese archivo "
            "en GitHub y hacé commit: la app redeploya y queda fijo. "
            "El uploader de acá abajo es solo un override temporal de la sesión."
        )
        if origen == "repo":
            st.success("AGENTE cargado desde el repo. Persistente, no hace falta subir nada.")
        elif origen == "sesion":
            st.info("Usando un AGENTE subido en esta sesión (se pierde al reiniciar). Para fijarlo, commiteálo en el repo.")
        elif origen == "disco":
            st.warning(
                "Hay una copia en disco del contenedor, pero en Streamlit Cloud "
                "es efímera. Committeá AGENTE_PUBLICADOR.xlsx en el repo para que sea permanente."
            )
        else:
            st.warning(
                "No hay AGENTE en el repo. Subilo acá para probar ahora, y "
                "committeá AGENTE_PUBLICADOR.xlsx en el repo para que quede fijo."
            )
        archivo_agente = st.file_uploader(
            "AGENTE_PUBLICADOR.xlsx (override de sesión)",
            type=["xlsx"],
            key="ap_agente_uploader"
        )
        if archivo_agente is not None:
            _guardar_agente_ap(archivo_agente.getvalue())
            st.success("AGENTE cargado para esta sesión.")

    if not _agente_existe():
        st.warning("Primero subí el AGENTE_PUBLICADOR en la sección de arriba.")
        return

    # Parsear AGENTE
    agente_bytes = _leer_agente_ap()

    try:
        bloques = parsear_agente_publicador_ap(agente_bytes)
    except Exception as e:
        st.error("Error al leer el AGENTE.")
        st.exception(e)
        return

    # --- Subir Publicar ---
    archivo_publicar = st.file_uploader(
        "Subí el Excel Publicar de Meli",
        type=["xlsx"],
        key="ap_publicar_uploader"
    )

    if archivo_publicar is None:
        return

    # Detectar categorías disponibles por hojas.
    # Solo leemos los nombres de hoja (read_only + cache): la carga completa
    # y escribible del Publicar se hace recién al apretar el botón.
    try:
        sheetnames_pub = _sheetnames_publicar_ap(archivo_publicar.getvalue())
    except Exception as e:
        st.error("No se pudo abrir el Publicar.")
        st.exception(e)
        return

    hojas_publicar = [h for h in sheetnames_pub if h not in ("Ayuda", "extra info")]

    # Matchear bloques del AGENTE con hojas del Publicar
    bloques_disponibles = []
    for b in bloques:
        hoja = encontrar_hoja_publicar(sheetnames_pub, b["nombre_limpio"])
        if hoja:
            b["hoja_publicar"] = hoja
            bloques_disponibles.append(b)

    if not bloques_disponibles:
        st.error(
            "No encontré ninguna categoría del AGENTE en las hojas del Publicar. "
            f"Hojas del Publicar: {hojas_publicar}"
        )
        return

    st.success(
        f"{len(bloques_disponibles)} categorías detectadas en el Publicar: "
        + ", ".join(b["etiqueta"] for b in bloques_disponibles)
    )

    # --- Seleccionar categoría ---
    etiquetas = [b["etiqueta"] for b in bloques_disponibles]
    seleccion = st.selectbox("Categoría a publicar", etiquetas, key="ap_cat_select")
    bloque = next(b for b in bloques_disponibles if b["etiqueta"] == seleccion)

    st.divider()

    # --- Campos fijos (informativos) ---
    with st.expander(f"Campos fijos ({len(bloque['campos_fijos'])} campos pre-cargados)"):
        for campo in bloque["campos_fijos"][:15]:
            st.caption(f"**{campo['header_norm']}**: {campo['valor']}")
        if len(bloque["campos_fijos"]) > 15:
            st.caption(f"... y {len(bloque['campos_fijos']) - 15} más")

    # --- Campos a solicitar ---
    st.subheader("Completá estos campos")

    valores_solicitados = {}
    campos_no_titulo = [c for c in bloque["campos_solicitar"] if not c["es_titulo"]]

    for campo in campos_no_titulo:
        label = campo["label"]
        tipo = campo.get("tipo", "texto")

        if tipo == "select_sino":
            valor = st.selectbox(
                label, ["Sí", "No"],
                key=f"ap_sol_{seleccion}_{label}",
            )
        elif "descripci" in label.lower():
            valor = st.text_area(
                label, height=150, key=f"ap_sol_{seleccion}_{label}",
                placeholder="Pegá la descripción completa"
            )
        elif "foto" in label.lower():
            valor = st.text_area(
                label, height=80, key=f"ap_sol_{seleccion}_{label}",
                placeholder="Links de fotos separados por coma"
            )
        else:
            valor = st.text_input(label, key=f"ap_sol_{seleccion}_{label}")

        valores_solicitados[label] = valor

    # --- Títulos ---
    st.subheader("Títulos (uno por línea)")

    titulos_texto = st.text_area(
        "Títulos",
        height=200,
        key=f"ap_titulos_{seleccion}",
        placeholder="Kit Seguridad Auto 7 En 1 Matafuego Baliza Chaleco\nKit Seguridad Vehicular Crique Hidraulico Llave Cruz\n..."
    )

    titulos = [t.strip() for t in titulos_texto.strip().split("\n") if t.strip()]

    if titulos:
        st.caption(f"{len(titulos)} títulos cargados.")

    # --- Generar ---
    if st.button("Completar Publicar", type="primary", key="ap_generar"):

        if not titulos:
            st.error("Cargá al menos un título.")
            return

        vacios = [l for l, v in valores_solicitados.items() if not str(v).strip()]
        if vacios:
            st.warning(f"Campos vacíos: {', '.join(vacios)}. Se dejan en blanco en el Excel.")

        try:
            # Recargar el archivo fresco para modificar
            wb = load_workbook(BytesIO(archivo_publicar.getvalue()))

            filas_escritas = escribir_publicaciones_ap(
                wb, bloque["hoja_publicar"], bloque, titulos, valores_solicitados
            )

            salida = BytesIO()
            wb.save(salida)
            salida.seek(0)

        except Exception as e:
            st.error("Error al completar el Publicar.")
            st.exception(e)
            return

        st.success(
            f"{filas_escritas} publicaciones escritas en la hoja "
            f"'{bloque['hoja_publicar']}' del Publicar."
        )

        c1, c2 = st.columns(2)
        c1.metric("Títulos", filas_escritas)
        c2.metric("Campos fijos por fila", len(bloque["campos_fijos"]))

        st.download_button(
            "Descargar Publicar completado",
            data=salida,
            file_name=archivo_publicar.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ap_descarga"
        )


# HERRAMIENTA 3 — BEST PRICE
# ============================================================
# Compara precios finales (con IVA) por artículo entre proveedores
# que pueden tener fórmulas de costo distintas.
#
# Config persistente en best_price_config.json:
# - registro de proveedores (fórmula, descuento, IVA, si detecta cajas)
# - overrides manuales de unidades por caja, por código
#
# NOTA: en Streamlit Community Cloud el filesystem es efímero.
# Si la app hiberna (12hs sin uso) y se reactiva, este JSON se
# pierde. Los defaults de Mansilla/Goicoechea se recargan solos;
# lo que se pierde son overrides manuales y proveedores agregados
# a mano. Migrar a Google Sheets si esto empieza a molestar.
# ============================================================

ARCHIVO_CONFIG_BEST_PRICE = "best_price_config.json"
CARPETA_PROVEEDORES_BP = "proveedores_data"

IVA_DEFAULT_BEST_PRICE = 21.0

# tipo_formula:
#   "descuento_iva"  -> final = base * (1 - descuento%) * (1 + iva%)
#   "lista_mas_iva"  -> final = base * (1 + iva%)   (sin descuento)
#
# modo_carga:
#   "generico"                -> una sola columna de precio (lista/precio/neto/costo)
#   "precio_condicional_oferta" -> usa "Precio Neto Oferta" si tiene valor > 0,
#                                   si no usa "Precio" (caso Iturria)
#
# detectar_cajas: si True, busca patrones de packs de aceites/fluidos
# en la descripción y divide por unidades. Si False, siempre 1 unidad.
PROVEEDORES_DEFAULT_BEST_PRICE = {
    "Mansilla": {
        "palabras_clave": ["mansilla"],
        "tipo_formula": "descuento_iva",
        "descuento": 38.0,
        "iva": 21.0,
        "detectar_cajas": True,
        "modo_carga": "generico",
    },
    "Goicoechea": {
        "palabras_clave": ["goicoechea"],
        "tipo_formula": "lista_mas_iva",
        "descuento": 0.0,
        "iva": 21.0,
        "detectar_cajas": False,
        "modo_carga": "generico",
    },
    "Iturria": {
        "palabras_clave": ["iturria"],
        "tipo_formula": "lista_mas_iva",
        "descuento": 0.0,
        "iva": 21.0,
        "detectar_cajas": False,
        "modo_carga": "precio_condicional_oferta",
    },
    "Pappiertei": {
        "palabras_clave": ["pappiertei"],
        "tipo_formula": "lista_mas_iva",
        "descuento": 0.0,
        "iva": 21.0,
        "detectar_cajas": False,
        "modo_carga": "generico",
    },
}

CODIGO_KEYWORDS_BP = ["codigo", "sku", "item_code", "cod", "pieza"]
# Orden = prioridad. "articulo" va al final porque en algunos proveedores
# (ej. Pappiertei) esa columna trae el fabricante, no la descripción real.
DESCRIPCION_KEYWORDS_BP = ["descripcion", "detalle", "articulo"]
PRECIO_KEYWORDS_BP = ["lista", "precio", "price"]
# Se usan SOLO si no aparece ninguna columna de lista/precio (ej: Goicoechea
# real -> "Neto"; Pappiertei -> "costo"). Queda marcado en la UI cuando se
# usa, para que quede claro que la base de cálculo es distinta.
PRECIO_KEYWORDS_FALLBACK_BP = ["neto", "costo"]

FILAS_ESCANEO_PROVEEDOR_BP = 15
FILAS_MAX_BUSQUEDA_HEADER_BP = 20

# Contexto que habilita interpretar "caja": solo fluidos/lubricantes.
REGEX_CONTEXTO_FLUIDO = (
    r"ACEITE|LIQ|FLUIDO|GRASA|ATF|HELICOIDAL|HIPOIDAL|DOT|REFRIGER|"
    r"DEXRON|DEXOS|LUBRICANTE|COOLANT|\d+W-?\d+"
)

# Palabras que descartan pack aunque haya patrón NxM
# (tornillería M8X1.25, llantas 17X7, tracción 4x4, etc.)
REGEX_EXCLUSION_PACK = (
    r"PERNO|TORNILLO|TUERCA|TAPON|LLANTA|RUEDA|SENSOR|BOMBA|"
    r"FILTRO|CARTER|TAPA|JUNTA|RETEN|SELLO|TUBO|VALVULA|"
    r"\bM\d+\s*[xX]|\b4\s*[xX]\s*4\b"
)


def cargar_config_best_price():
    config = {
        "proveedores": {
            nombre: dict(datos)
            for nombre, datos in PROVEEDORES_DEFAULT_BEST_PRICE.items()
        },
        "unidades_override": {},
    }

    if os.path.exists(ARCHIVO_CONFIG_BEST_PRICE):
        try:
            with open(ARCHIVO_CONFIG_BEST_PRICE, "r", encoding="utf-8") as f:
                guardado = json.load(f)

            if isinstance(guardado.get("proveedores"), dict):
                for nombre, datos in guardado["proveedores"].items():
                    if nombre in config["proveedores"]:
                        config["proveedores"][nombre].update(datos)
                    else:
                        datos.setdefault("modo_carga", "generico")
                        config["proveedores"][nombre] = datos

            if isinstance(guardado.get("unidades_override"), dict):
                config["unidades_override"] = {
                    str(k): int(v)
                    for k, v in guardado["unidades_override"].items()
                    if str(v).strip() != ""
                }

        except Exception:
            pass

    return config


def guardar_config_best_price(config):
    try:
        with open(ARCHIVO_CONFIG_BEST_PRICE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


# ------------------------------------------------------------
# Persistencia de archivos de proveedores en disco
# ------------------------------------------------------------

def asegurar_carpeta_proveedores():
    os.makedirs(CARPETA_PROVEEDORES_BP, exist_ok=True)


def ruta_archivo_proveedor(nombre_proveedor):
    return os.path.join(CARPETA_PROVEEDORES_BP, f"{nombre_proveedor}.xlsx")


def guardar_archivo_proveedor(nombre_proveedor, archivo_bytes):
    asegurar_carpeta_proveedores()
    with open(ruta_archivo_proveedor(nombre_proveedor), "wb") as f:
        f.write(archivo_bytes)


def eliminar_archivo_proveedor(nombre_proveedor):
    ruta = ruta_archivo_proveedor(nombre_proveedor)
    if os.path.exists(ruta):
        os.remove(ruta)


def listar_proveedores_activos():
    asegurar_carpeta_proveedores()
    activos = {}

    for archivo in os.listdir(CARPETA_PROVEEDORES_BP):
        if not archivo.lower().endswith(".xlsx"):
            continue

        nombre_proveedor = archivo[:-5]
        ruta = os.path.join(CARPETA_PROVEEDORES_BP, archivo)

        try:
            stat = os.stat(ruta)
            activos[nombre_proveedor] = {
                "ruta": ruta,
                "modificado": datetime.fromtimestamp(stat.st_mtime),
                "tamano_kb": round(stat.st_size / 1024, 1),
            }
        except Exception:
            continue

    return activos


def detectar_proveedor_por_nombre_archivo(nombre_archivo, registro_proveedores):
    base = normalizar_texto(os.path.splitext(nombre_archivo)[0])

    for nombre_prov, cfg in registro_proveedores.items():
        palabras = cfg.get("palabras_clave") or [normalizar_texto(nombre_prov)]

        for palabra in palabras:
            if normalizar_texto(palabra) and normalizar_texto(palabra) in base:
                return nombre_prov

    return None


def detectar_proveedor_en_archivo(archivo_bytes, registro_proveedores):
    try:
        wb = load_workbook(BytesIO(archivo_bytes), read_only=True, data_only=True)
    except Exception:
        return None

    textos = []

    for ws in wb.worksheets:
        for row in ws.iter_rows(max_row=FILAS_ESCANEO_PROVEEDOR_BP):
            for cell in row:
                if cell.value is not None:
                    textos.append(normalizar_texto(cell.value))
        break

    texto_completo = " ".join(textos)

    for nombre_prov, cfg in registro_proveedores.items():
        palabras = cfg.get("palabras_clave") or [normalizar_texto(nombre_prov)]

        for palabra in palabras:
            if normalizar_texto(palabra) and normalizar_texto(palabra) in texto_completo:
                return nombre_prov

    return None


def resolver_proveedor_de_archivo(nombre_archivo, archivo_bytes, config, key_prefix):
    detectado = detectar_proveedor_por_nombre_archivo(nombre_archivo, config["proveedores"])
    metodo = "nombre del archivo"

    if detectado is None:
        detectado = detectar_proveedor_en_archivo(archivo_bytes, config["proveedores"])
        metodo = "contenido del archivo"

    if detectado is not None:
        st.success(f"**{nombre_archivo}** → detectado como **{detectado}** (por {metodo})")
        return detectado

    st.warning(f"**{nombre_archivo}**: no pude reconocer el proveedor. Elegilo:")

    return st.selectbox(
        "Proveedor",
        sorted(config["proveedores"].keys()),
        key=f"{key_prefix}_manual_{nombre_archivo}"
    )


def encontrar_fila_encabezado_bp(archivo_bytes, hoja):
    wb = load_workbook(BytesIO(archivo_bytes), read_only=True, data_only=True)
    ws = wb[hoja]

    mejor_fila = None
    mejor_score = 0

    for i, row in enumerate(
        ws.iter_rows(max_row=FILAS_MAX_BUSQUEDA_HEADER_BP, values_only=True), start=1
    ):
        tiene_codigo = tiene_desc = tiene_precio = False

        for valor in row:
            nombre = normalizar_texto(valor)
            if not nombre:
                continue
            if not tiene_codigo and any(k in nombre for k in CODIGO_KEYWORDS_BP):
                tiene_codigo = True
            if not tiene_desc and any(k in nombre for k in DESCRIPCION_KEYWORDS_BP):
                tiene_desc = True
            if not tiene_precio and any(
                k in nombre for k in PRECIO_KEYWORDS_BP + PRECIO_KEYWORDS_FALLBACK_BP
            ):
                tiene_precio = True

        score = int(tiene_codigo) + int(tiene_desc) + int(tiene_precio)

        if score >= 2 and score > mejor_score:
            mejor_score = score
            mejor_fila = i

    return mejor_fila


def _buscar_columna_por_prioridad(columnas, keywords, ya_asignadas):
    """
    Recorre KEYWORDS en orden de prioridad; para cada uno, busca la
    primera columna (no asignada todavía a otro campo) que lo contenga.
    Evita que una columna genérica (ej. "articulo") le gane a una más
    específica (ej. "detalle") solo por venir antes en el archivo.
    """
    for kw in keywords:
        for col in columnas:
            if col in ya_asignadas:
                continue
            if kw in normalizar_texto(col):
                return col
    return None


def resolver_columnas_lista_bp(columnas):
    """
    Devuelve (col_codigo, col_descripcion, col_precio, precio_es_fallback).
    precio_es_fallback = True cuando no había columna de "lista"/"precio"
    y se usó "neto" o "costo" en su lugar.
    """
    columnas = list(columnas)
    ya_asignadas = set()

    col_codigo = _buscar_columna_por_prioridad(columnas, CODIGO_KEYWORDS_BP, ya_asignadas)
    if col_codigo:
        ya_asignadas.add(col_codigo)

    col_descripcion = _buscar_columna_por_prioridad(columnas, DESCRIPCION_KEYWORDS_BP, ya_asignadas)
    if col_descripcion:
        ya_asignadas.add(col_descripcion)

    col_precio = _buscar_columna_por_prioridad(columnas, PRECIO_KEYWORDS_BP, ya_asignadas)
    precio_es_fallback = False

    if col_precio is None:
        col_precio = _buscar_columna_por_prioridad(columnas, PRECIO_KEYWORDS_FALLBACK_BP, ya_asignadas)
        precio_es_fallback = col_precio is not None

    faltantes = []
    if col_codigo is None:
        faltantes.append("código")
    if col_descripcion is None:
        faltantes.append("descripción")
    if col_precio is None:
        faltantes.append("precio (ni lista/precio ni neto/costo)")

    if faltantes:
        raise ValueError(
            "No se detectaron estas columnas: " + ", ".join(faltantes)
            + ". Columnas disponibles: " + ", ".join(map(str, columnas))
        )

    return col_codigo, col_descripcion, col_precio, precio_es_fallback


def resolver_columnas_iturria_bp(columnas):
    """
    Caso especial: el costo depende de dos columnas ("Precio" y
    "Precio Neto Oferta"), no de una sola.
    """
    columnas = list(columnas)

    col_codigo = _buscar_columna_por_prioridad(columnas, CODIGO_KEYWORDS_BP, set())
    col_descripcion = _buscar_columna_por_prioridad(columnas, DESCRIPCION_KEYWORDS_BP, set())

    col_precio_base = None
    for col in columnas:
        n = normalizar_texto(col)
        if "precio" in n and "oferta" not in n and "neto" not in n and "catalogo" not in n:
            col_precio_base = col
            break

    col_precio_oferta_neto = None
    for col in columnas:
        n = normalizar_texto(col)
        if "neto" in n and "oferta" in n:
            col_precio_oferta_neto = col
            break

    faltantes = []
    if col_codigo is None:
        faltantes.append("código")
    if col_descripcion is None:
        faltantes.append("descripción")
    if col_precio_base is None:
        faltantes.append("columna 'Precio'")
    if col_precio_oferta_neto is None:
        faltantes.append("columna 'Precio Neto Oferta'")

    if faltantes:
        raise ValueError(
            "No se detectaron estas columnas: " + ", ".join(faltantes)
            + ". Columnas disponibles: " + ", ".join(map(str, columnas))
        )

    return col_codigo, col_descripcion, col_precio_base, col_precio_oferta_neto


@st.cache_data(show_spinner=False)
def cargar_lista_generica_bp(archivo_bytes):
    """
    Lee el Excel de un proveedor con una sola columna de precio,
    sin asumir en qué fila está el encabezado. Devuelve
    (DataFrame[CODIGO, DESCRIPCION, PRECIO_LISTA], precio_es_fallback).
    """
    archivo_bytes = reparar_tablas_duplicadas_excel(archivo_bytes)

    xls = pd.ExcelFile(BytesIO(archivo_bytes))
    frames = []
    fallback_usado = False

    for hoja in xls.sheet_names:
        if str(hoja).upper().startswith("CONTROL"):
            continue

        fila_header = encontrar_fila_encabezado_bp(archivo_bytes, hoja)

        if fila_header is None:
            continue

        df = pd.read_excel(
            BytesIO(archivo_bytes),
            sheet_name=hoja,
            header=fila_header - 1,
            dtype=str
        )

        if df.empty:
            continue

        try:
            col_cod, col_desc, col_precio, es_fallback = resolver_columnas_lista_bp(df.columns)
        except ValueError:
            continue

        if es_fallback:
            fallback_usado = True

        parcial = pd.DataFrame({
            "CODIGO": df[col_cod].apply(normalizar_sku),
            "DESCRIPCION": df[col_desc].fillna("").astype(str).str.strip(),
            "PRECIO_LISTA": df[col_precio].apply(convertir_precio_a_numero),
        })

        parcial = parcial[(parcial["CODIGO"] != "") & parcial["PRECIO_LISTA"].notna()]
        frames.append(parcial)

    if not frames:
        raise ValueError(
            "No se encontraron columnas de código, descripción y precio "
            "reconocibles en ninguna hoja del archivo."
        )

    base = pd.concat(frames, ignore_index=True)
    base = base.drop_duplicates(subset="CODIGO", keep="first").reset_index(drop=True)

    return base, fallback_usado


@st.cache_data(show_spinner=False)
def cargar_lista_iturria_bp(archivo_bytes):
    """
    Caso Iturria: PRECIO_LISTA = Precio Neto Oferta si tiene valor > 0,
    si no, Precio (ambos sin IVA, que se suma después según config).
    Devuelve (DataFrame[CODIGO, DESCRIPCION, PRECIO_LISTA, FUENTE_PRECIO], False).
    El segundo valor (fallback) siempre False acá, se mantiene por
    compatibilidad de firma con el loader genérico.
    """
    archivo_bytes = reparar_tablas_duplicadas_excel(archivo_bytes)

    xls = pd.ExcelFile(BytesIO(archivo_bytes))
    frames = []

    for hoja in xls.sheet_names:
        if str(hoja).upper().startswith("CONTROL"):
            continue

        fila_header = encontrar_fila_encabezado_bp(archivo_bytes, hoja)

        if fila_header is None:
            continue

        df = pd.read_excel(
            BytesIO(archivo_bytes),
            sheet_name=hoja,
            header=fila_header - 1,
            dtype=str
        )

        if df.empty:
            continue

        try:
            col_cod, col_desc, col_base, col_oferta = resolver_columnas_iturria_bp(df.columns)
        except ValueError:
            continue

        precio_base = df[col_base].apply(convertir_precio_a_numero)
        precio_oferta = df[col_oferta].apply(convertir_precio_a_numero)

        precio_final = []
        fuente = []

        for base_val, oferta_val in zip(precio_base, precio_oferta):
            if oferta_val is not None and oferta_val > 0:
                precio_final.append(oferta_val)
                fuente.append("Precio Neto Oferta")
            else:
                precio_final.append(base_val)
                fuente.append("Precio")

        parcial = pd.DataFrame({
            "CODIGO": df[col_cod].apply(normalizar_sku),
            "DESCRIPCION": df[col_desc].fillna("").astype(str).str.strip(),
            "PRECIO_LISTA": precio_final,
            "FUENTE_PRECIO": fuente,
        })

        parcial = parcial[(parcial["CODIGO"] != "") & parcial["PRECIO_LISTA"].notna()]
        frames.append(parcial)

    if not frames:
        raise ValueError(
            "No se encontraron las columnas de Iturria (código, descripción, "
            "Precio, Precio Neto Oferta) en ninguna hoja del archivo."
        )

    base = pd.concat(frames, ignore_index=True)
    base = base.drop_duplicates(subset="CODIGO", keep="first").reset_index(drop=True)

    return base, False


def cargar_lista_proveedor_bp(archivo_bytes, cfg_prov):
    """
    Despacha al loader correcto según modo_carga del proveedor.
    """
    modo = cfg_prov.get("modo_carga", "generico")

    if modo == "precio_condicional_oferta":
        return cargar_lista_iturria_bp(archivo_bytes)

    return cargar_lista_generica_bp(archivo_bytes)


def detectar_unidades_pack_bp(descripcion):
    if descripcion is None:
        return None, None

    texto = str(descripcion).upper()

    if not re.search(REGEX_CONTEXTO_FLUIDO, texto):
        return None, None

    if re.search(REGEX_EXCLUSION_PACK, texto):
        return None, None

    m = re.search(r"\b(\d{1,2})\s*[xX]\s*\d+(?:[.,]\d+)?\s*(?:LITROS?|LTS?|L)\b", texto)
    if m and int(m.group(1)) >= 2:
        return int(m.group(1)), "alta"

    m = re.search(r"CAJA\s*(\d{1,2})\s*[xX]", texto)
    if m and int(m.group(1)) >= 2:
        return int(m.group(1)), "alta"

    m = re.search(r"GM\s*(\d{1,2})-(\d{1,4})\b", texto)
    if m and int(m.group(1)) >= 2:
        return int(m.group(1)), "alta"

    m = re.search(r"\b(\d{1,2})\s*[xX]\s*\d+(?:[.,]\d+)?\s*$", texto)
    if m and int(m.group(1)) >= 2:
        return int(m.group(1)), "verificar"

    m = re.search(r"\bX\s*(\d{1,2})\s*$", texto)
    if m and int(m.group(1)) >= 2:
        return int(m.group(1)), "verificar"

    m = re.search(r"\s-\s*(\d{1,2})(?:\s*[xX])?\s*$", texto)
    if m and int(m.group(1)) in (2, 3, 4, 6, 12, 24):
        return int(m.group(1)), "verificar"

    return None, None


def calcular_precios_proveedor(base, cfg, overrides):
    """
    Aplica la fórmula del proveedor y, si corresponde, la división
    por unidades de caja. Siempre devuelve PRECIO_FINAL_UNIDAD,
    que es la columna que se usa para comparar entre proveedores.
    """
    resultado = base.copy()
    tipo = cfg["tipo_formula"]
    detecta_cajas = tipo == "descuento_iva" and cfg.get("detectar_cajas", False)

    if detecta_cajas:
        deteccion = resultado["DESCRIPCION"].apply(detectar_unidades_pack_bp)
        resultado["UNID_DETECTADAS"] = [d[0] for d in deteccion]
        resultado["DETECCION"] = [d[1] for d in deteccion]
    else:
        resultado["UNID_DETECTADAS"] = None
        resultado["DETECCION"] = None

    unidades = []
    origen_unidades = []

    for _, fila in resultado.iterrows():
        codigo = str(fila["CODIGO"])

        if not detecta_cajas:
            unidades.append(1)
            origen_unidades.append("unidad")
        elif codigo in overrides:
            unidades.append(max(1, int(overrides[codigo])))
            origen_unidades.append("manual")
        elif fila["UNID_DETECTADAS"] is not None and not pd.isna(fila["UNID_DETECTADAS"]):
            unidades.append(int(fila["UNID_DETECTADAS"]))
            origen_unidades.append(
                "auto" if fila["DETECCION"] == "alta" else "auto (verificar)"
            )
        else:
            unidades.append(1)
            origen_unidades.append("unidad")

    resultado["UNID_X_CAJA"] = unidades
    resultado["ORIGEN_UNIDADES"] = origen_unidades

    if tipo == "descuento_iva":
        factor = (1 - cfg.get("descuento", 0.0) / 100.0) * (1 + cfg["iva"] / 100.0)
    elif tipo == "lista_mas_iva":
        factor = 1 + cfg["iva"] / 100.0
    else:
        raise ValueError(f"tipo_formula desconocido: {tipo}")

    resultado["PRECIO_FINAL_CAJA"] = (resultado["PRECIO_LISTA"] * factor).round(2)
    resultado["PRECIO_FINAL_UNIDAD"] = (
        resultado["PRECIO_FINAL_CAJA"] / resultado["UNID_X_CAJA"]
    ).round(2)

    return resultado


def armar_tabla_comparacion(resultados_por_proveedor):
    nombres = list(resultados_por_proveedor.keys())

    descripciones = pd.concat(
        [df.set_index("CODIGO")["DESCRIPCION"] for df in resultados_por_proveedor.values()]
    )
    descripciones = descripciones[~descripciones.index.duplicated(keep="first")]

    combinado = descripciones.to_frame("Descripción")

    for nombre in nombres:
        df = resultados_por_proveedor[nombre]
        serie = df.set_index("CODIGO")["PRECIO_FINAL_UNIDAD"]
        combinado[nombre] = serie

    combinado["Proveedor más barato"] = combinado[nombres].idxmin(axis=1, skipna=True)
    combinado["Precio final más barato"] = combinado[nombres].min(axis=1, skipna=True)

    combinado = combinado.sort_values(
        "Precio final más barato", ascending=True, na_position="last"
    )

    combinado = combinado.reset_index().rename(columns={"CODIGO": "Código"})

    return combinado


def mostrar_configuracion_proveedores_bp(config):
    with st.expander("Configuración de proveedores"):

        st.caption(
            "Los valores quedan guardados hasta que los cambies. "
            "Si la app hiberna y se reactiva en Streamlit Cloud, "
            "los defaults se recargan solos."
        )

        nombres = sorted(config["proveedores"].keys())

        proveedor_editar = st.selectbox(
            "Proveedor a editar", nombres, key="bp_proveedor_config"
        )

        datos_prov = config["proveedores"][proveedor_editar]
        modo_carga_prov = datos_prov.get("modo_carga", "generico")

        if modo_carga_prov == "precio_condicional_oferta":
            st.caption(
                "Este proveedor usa una regla especial: toma 'Precio Neto Oferta' "
                "si tiene valor, y si no, 'Precio'. No se puede editar desde acá."
            )

        tipo_actual = datos_prov.get("tipo_formula", "descuento_iva")

        tipo_formula = st.radio(
            "Cómo se calcula el costo",
            options=["descuento_iva", "lista_mas_iva"],
            format_func=lambda v: (
                "Base − descuento% + IVA" if v == "descuento_iva"
                else "Base + IVA, sin descuento"
            ),
            index=0 if tipo_actual == "descuento_iva" else 1,
            key=f"bp_tipo_formula_{proveedor_editar}",
            horizontal=True,
        )

        c1, c2, c3 = st.columns(3)

        with c1:
            if tipo_formula == "descuento_iva":
                nuevo_descuento = st.number_input(
                    "Descuento sobre base (%)",
                    min_value=0.0, max_value=99.0,
                    value=float(datos_prov.get("descuento", 0.0)),
                    step=0.5,
                    key=f"bp_descuento_{proveedor_editar}"
                )
            else:
                nuevo_descuento = 0.0
                st.caption("Sin descuento.")

        with c2:
            nuevo_iva = st.number_input(
                "IVA (%)",
                min_value=0.0, max_value=50.0,
                value=float(datos_prov.get("iva", IVA_DEFAULT_BEST_PRICE)),
                step=0.5,
                key=f"bp_iva_{proveedor_editar}"
            )

        with c3:
            if tipo_formula == "descuento_iva" and modo_carga_prov == "generico":
                nuevo_detectar_cajas = st.checkbox(
                    "Detectar aceites por caja",
                    value=bool(datos_prov.get("detectar_cajas", False)),
                    key=f"bp_detcajas_{proveedor_editar}",
                    help="Si algunos artículos (aceites/fluidos) se venden por caja."
                )
            else:
                nuevo_detectar_cajas = False
                st.caption("No aplica para este proveedor.")

        palabras_actuales = ", ".join(datos_prov.get("palabras_clave", [proveedor_editar]))
        nuevas_palabras = st.text_input(
            "Palabra(s) clave para reconocer a este proveedor "
            "(separadas por coma; se busca primero en el nombre del archivo, "
            "y si no aparece ahí, dentro del contenido del Excel)",
            value=palabras_actuales,
            key=f"bp_palabras_{proveedor_editar}"
        )

        if st.button("Guardar configuración", key="bp_guardar_config"):
            config["proveedores"][proveedor_editar]["tipo_formula"] = tipo_formula
            config["proveedores"][proveedor_editar]["descuento"] = nuevo_descuento
            config["proveedores"][proveedor_editar]["iva"] = nuevo_iva
            config["proveedores"][proveedor_editar]["detectar_cajas"] = nuevo_detectar_cajas
            config["proveedores"][proveedor_editar]["palabras_clave"] = [
                p.strip() for p in nuevas_palabras.split(",") if p.strip()
            ] or [proveedor_editar]
            # modo_carga no se toca desde la UI, queda como estaba

            if guardar_config_best_price(config):
                st.success(f"Configuración de {proveedor_editar} guardada.")
            else:
                st.error("No se pudo escribir best_price_config.json.")

        st.divider()
        st.markdown("**Agregar proveedor nuevo**")
        st.caption(
            "Los proveedores nuevos usan el modo de carga genérico (una sola "
            "columna de precio). Si necesitás una regla especial como la de "
            "Iturria, avisame para agregarla a mano."
        )

        cn1, cn2 = st.columns(2)
        with cn1:
            nombre_nuevo = st.text_input("Nombre", key="bp_nuevo_nombre")
        with cn2:
            tipo_nuevo = st.radio(
                "Fórmula", options=["descuento_iva", "lista_mas_iva"],
                format_func=lambda v: (
                    "Base − descuento% + IVA" if v == "descuento_iva"
                    else "Base + IVA"
                ),
                key="bp_nuevo_tipo", horizontal=True,
            )

        if st.button("Agregar proveedor", key="bp_agregar_proveedor"):
            nombre_limpio = str(nombre_nuevo).strip()

            if nombre_limpio == "":
                st.error("El nombre del proveedor no puede estar vacío.")
            elif nombre_limpio in config["proveedores"]:
                st.error("Ese proveedor ya existe. Editalo arriba.")
            else:
                config["proveedores"][nombre_limpio] = {
                    "tipo_formula": tipo_nuevo,
                    "descuento": 0.0,
                    "iva": IVA_DEFAULT_BEST_PRICE,
                    "detectar_cajas": False,
                    "modo_carga": "generico",
                    "palabras_clave": [normalizar_texto(nombre_limpio)],
                }
                guardar_config_best_price(config)
                st.success(f"Proveedor {nombre_limpio} agregado. Configurá sus valores arriba.")
                st.rerun()


def mostrar_gestor_proveedores_activos(config):
    st.subheader("Proveedores activos")

    activos = listar_proveedores_activos()

    if activos:
        for nombre_prov in sorted(activos.keys()):
            info = activos[nombre_prov]
            c1, c2, c3, c4 = st.columns([2, 2, 2, 1])

            c1.markdown(f"**{nombre_prov}**")
            c2.caption(f"Actualizado: {info['modificado'].strftime('%d/%m/%Y %H:%M')}")
            c3.caption(f"{info['tamano_kb']} KB")

            if c4.button("Quitar", key=f"bp_quitar_{nombre_prov}"):
                eliminar_archivo_proveedor(nombre_prov)
                st.rerun()
    else:
        st.caption("Todavía no subiste ningún archivo de proveedor.")

    st.markdown("**Cargar / reemplazar archivo de un proveedor**")
    st.caption(
        "Nombrá el archivo igual que el proveedor (ej: Mansilla.xlsx, "
        "Goicoechea.xlsx, Iturria.xlsx, Pappiertei.xlsx) para que lo "
        "identifique solo. Si subís uno con el mismo proveedor, reemplaza "
        "al anterior; los demás no se tocan."
    )

    nuevo_archivo = st.file_uploader(
        "Subir archivo de proveedor",
        type=["xlsx"],
        key="bp_uploader_activo"
    )

    if nuevo_archivo is not None:
        proveedor = resolver_proveedor_de_archivo(
            nuevo_archivo.name, nuevo_archivo.getvalue(), config, "bp_activo"
        )

        if st.button(f"Guardar como archivo activo de {proveedor}", key="bp_confirmar_guardado"):
            guardar_archivo_proveedor(proveedor, nuevo_archivo.getvalue())
            st.success(f"{proveedor}: archivo guardado como activo.")
            st.rerun()

    return listar_proveedores_activos()


def mostrar_modo_comparacion_bp(config, activos):
    st.subheader("Comparar proveedores")
    st.caption(
        "Usa los archivos activos de arriba, cruzando por código de pieza. "
        "El ranking completo sin buscar mezcla catálogos y fichas técnicas a "
        "precio simbólico — usalo siempre con el buscador."
    )

    if len(activos) < 1:
        st.info("Subí al menos un archivo de proveedor arriba para poder comparar.")
        return

    resultados_por_proveedor = {}

    for nombre_prov, info in activos.items():
        cfg_prov = config["proveedores"].get(nombre_prov)

        if cfg_prov is None:
            st.warning(
                f"{nombre_prov} no tiene configuración de fórmula. "
                "Agregalo en 'Configuración de proveedores'."
            )
            continue

        with open(info["ruta"], "rb") as f:
            archivo_bytes = f.read()

        try:
            base, fallback_usado = cargar_lista_proveedor_bp(archivo_bytes, cfg_prov)
        except Exception as e:
            st.error(f"No se pudo leer el archivo activo de {nombre_prov}.")
            st.exception(e)
            continue

        if fallback_usado:
            st.caption(
                f"ℹ️ {nombre_prov}: no había columna de lista/precio, "
                "se usó 'Neto'/'Costo' como base de cálculo."
            )

        resultados_por_proveedor[nombre_prov] = calcular_precios_proveedor(
            base, cfg_prov, config["unidades_override"]
        )

    if not resultados_por_proveedor:
        return

    m1, m2 = st.columns(2)
    m1.metric("Proveedores cargados", len(resultados_por_proveedor))
    m2.metric(
        "Artículos totales (únicos por código)",
        f"{sum(len(df) for df in resultados_por_proveedor.values()):,}".replace(",", ".")
    )

    comparacion = armar_tabla_comparacion(resultados_por_proveedor)

    busqueda = st.text_input(
        "Buscar por código o descripción",
        key="bp_busqueda_comparacion",
        placeholder="Ej: 5w30 dexos | 101996 | luneta captiva"
    )

    if busqueda.strip():
        terminos = [normalizar_texto(t) for t in busqueda.split() if t.strip()]

        texto_busqueda = (
            comparacion["Código"].astype(str) + " " + comparacion["Descripción"]
        ).apply(normalizar_texto)

        mascara = pd.Series(True, index=comparacion.index)
        for termino in terminos:
            mascara &= texto_busqueda.str.contains(re.escape(termino), na=False)

        comparacion = comparacion[mascara]

    if comparacion.empty:
        st.info("No hay artículos que coincidan con la búsqueda.")
        return

    LIMITE_FILAS_BP = 300
    total = len(comparacion)

    if total > LIMITE_FILAS_BP:
        st.caption(
            f"{total:,} coincidencias. Se muestran las primeras {LIMITE_FILAS_BP} "
            "(ya ordenadas por más barato). Afiná la búsqueda para acotar."
        )
        comparacion = comparacion.head(LIMITE_FILAS_BP)

    config_columnas = {
        nombre: st.column_config.NumberColumn(format="$ %.2f")
        for nombre in resultados_por_proveedor.keys()
    }
    config_columnas["Precio final más barato"] = st.column_config.NumberColumn(format="$ %.2f")

    st.dataframe(
        comparacion,
        use_container_width=True,
        hide_index=True,
        column_config=config_columnas,
    )

    salida = BytesIO()
    comparacion.to_excel(salida, index=False)
    salida.seek(0)

    st.download_button(
        "Descargar comparación (xlsx)",
        data=salida,
        file_name="best_price_comparacion.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="bp_descarga_comparacion"
    )


def mostrar_modo_detalle_bp(config, activos):
    st.subheader("Ver un proveedor")

    if not activos:
        st.info("Subí al menos un archivo de proveedor arriba.")
        return

    proveedor = st.selectbox(
        "Proveedor a ver",
        sorted(activos.keys()),
        key="bp_detalle_proveedor"
    )

    cfg_prov = config["proveedores"].get(proveedor)

    if cfg_prov is None:
        st.warning(f"{proveedor} no tiene configuración de fórmula todavía.")
        return

    with open(activos[proveedor]["ruta"], "rb") as f:
        archivo_bytes = f.read()

    if cfg_prov.get("modo_carga") == "precio_condicional_oferta":
        st.info(
            f"Fórmula {proveedor}: usa 'Precio Neto Oferta' si tiene valor, "
            f"si no usa 'Precio'; a ese valor se le suma IVA {cfg_prov['iva']}%. "
            "Siempre por unidad."
        )
    else:
        st.info(
            (
                f"Fórmula {proveedor}: base − {cfg_prov.get('descuento', 0)}% + IVA {cfg_prov['iva']}%."
                if cfg_prov["tipo_formula"] == "descuento_iva"
                else f"Fórmula {proveedor}: base + IVA {cfg_prov['iva']}% (sin descuento)."
            )
            + (
                " Si el artículo viene por caja, se divide por las unidades."
                if cfg_prov.get("detectar_cajas")
                else " Este proveedor siempre cotiza por unidad."
            )
        )

    try:
        with st.spinner("Leyendo lista de precios..."):
            base, fallback_usado = cargar_lista_proveedor_bp(archivo_bytes, cfg_prov)
    except Exception as e:
        st.error("No se pudo leer la lista del proveedor.")
        st.exception(e)
        return

    if fallback_usado:
        st.caption(
            "ℹ️ No había columna de lista/precio en este archivo, "
            "se usó 'Neto'/'Costo' como base de cálculo."
        )

    resultado = calcular_precios_proveedor(base, cfg_prov, config["unidades_override"])

    m1, m2, m3 = st.columns(3)
    m1.metric("Artículos en lista", f"{len(base):,}".replace(",", "."))

    if cfg_prov.get("detectar_cajas"):
        packs_detectados = int(resultado["UNID_DETECTADAS"].notna().sum())
        m2.metric("Cajas detectadas", packs_detectados)
    else:
        m2.metric("Cajas detectadas", "N/A (por unidad)")

    m3.metric("Unidades corregidas a mano", len(config["unidades_override"]))

    busqueda = st.text_input(
        "Buscar por código o descripción",
        key="bp_busqueda_detalle",
        placeholder="Ej: 5w30 dexos | 106396 | liq frenos dot 4"
    )

    df_filtrado = resultado

    if busqueda.strip():
        terminos = [normalizar_texto(t) for t in busqueda.split() if t.strip()]

        texto_busqueda = (
            df_filtrado["CODIGO"].astype(str) + " " + df_filtrado["DESCRIPCION"]
        ).apply(normalizar_texto)

        mascara = pd.Series(True, index=df_filtrado.index)
        for termino in terminos:
            mascara &= texto_busqueda.str.contains(re.escape(termino), na=False)

        df_filtrado = df_filtrado[mascara]

    if df_filtrado.empty:
        st.info("No hay artículos que coincidan con la búsqueda.")
        return

    LIMITE_FILAS_BP = 300
    total = len(df_filtrado)

    if total > LIMITE_FILAS_BP:
        st.caption(f"{total:,} coincidencias. Se muestran las primeras {LIMITE_FILAS_BP}.")
        df_filtrado = df_filtrado.head(LIMITE_FILAS_BP)

    col_final_caja = f"Final c/IVA {proveedor}"
    col_final_unidad = f"Final c/IVA x unidad {proveedor}"

    columnas_mostrar = {
        "CODIGO": "Código",
        "DESCRIPCION": "Descripción",
        "PRECIO_LISTA": "Precio base s/IVA",
        "PRECIO_FINAL_CAJA": col_final_caja,
        "PRECIO_FINAL_UNIDAD": col_final_unidad,
    }

    if "FUENTE_PRECIO" in df_filtrado.columns:
        columnas_mostrar["FUENTE_PRECIO"] = "Columna usada"

    if cfg_prov.get("detectar_cajas"):
        columnas_mostrar["UNID_X_CAJA"] = "Unid x Caja"
        columnas_mostrar["ORIGEN_UNIDADES"] = "Detección"

    tabla = df_filtrado[list(columnas_mostrar.keys())].rename(columns=columnas_mostrar)

    if cfg_prov.get("detectar_cajas"):
        editada = st.data_editor(
            tabla,
            use_container_width=True,
            hide_index=True,
            key="bp_editor_detalle",
            disabled=[c for c in tabla.columns if c != "Unid x Caja"],
            column_config={
                "Unid x Caja": st.column_config.NumberColumn(
                    min_value=1, max_value=48, step=1,
                    help="Editá si la caja tiene otra cantidad. Se guarda por código."
                ),
                "Precio base s/IVA": st.column_config.NumberColumn(format="$ %.2f"),
                col_final_caja: st.column_config.NumberColumn(format="$ %.2f"),
                col_final_unidad: st.column_config.NumberColumn(format="$ %.2f"),
            },
        )

        unidades_previas = tabla.set_index("Código")["Unid x Caja"].to_dict()
        hubo_cambios = False

        for _, fila in editada.iterrows():
            codigo = str(fila["Código"])
            try:
                unidades_nuevas = int(fila["Unid x Caja"])
            except Exception:
                continue

            if unidades_nuevas < 1:
                continue

            if unidades_nuevas != int(unidades_previas.get(codigo, 1)):
                config["unidades_override"][codigo] = unidades_nuevas
                hubo_cambios = True

        if hubo_cambios:
            guardar_config_best_price(config)
            st.rerun()
    else:
        column_config_detalle = {
            "Precio base s/IVA": st.column_config.NumberColumn(format="$ %.2f"),
            col_final_caja: st.column_config.NumberColumn(format="$ %.2f"),
            col_final_unidad: st.column_config.NumberColumn(format="$ %.2f"),
        }
        st.dataframe(
            tabla,
            use_container_width=True,
            hide_index=True,
            column_config=column_config_detalle,
        )

    salida = BytesIO()
    tabla.to_excel(salida, index=False)
    salida.seek(0)

    st.download_button(
        "Descargar resultado (xlsx)",
        data=salida,
        file_name=f"best_price_{normalizar_texto(proveedor).replace(' ', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="bp_descarga_detalle"
    )


def mostrar_herramienta_best_price():
    st.title("Best Price")
    st.caption(
        "Precio final con IVA por artículo. Subís cada lista una vez; queda "
        "activa hasta que la reemplaces. Cada proveedor puede tener su "
        "propia fórmula de costo."
    )

    config = cargar_config_best_price()

    mostrar_configuracion_proveedores_bp(config)

    st.divider()

    activos = mostrar_gestor_proveedores_activos(config)

    st.divider()

    modo = st.radio(
        "¿Qué querés hacer?",
        options=["comparar", "detalle"],
        format_func=lambda v: (
            "Comparar proveedores" if v == "comparar"
            else "Ver el detalle de un solo proveedor"
        ),
        key="bp_modo",
        horizontal=True,
    )

    st.divider()

    if modo == "comparar":
        mostrar_modo_comparacion_bp(config, activos)
    else:
        mostrar_modo_detalle_bp(config, activos)


# ============================================================
# HERRAMIENTA 4 — ANALISTA GENERAL (FASE 1)
# ============================================================
# Fase 1: 3 tableros independientes, cada uno con su propio archivo.
# Todavía NO cruza margen interno (por pieza) contra performance de
# Meli (por kit/publicación) porque usan claves de producto distintas
# -- ver aclaración en el chat. Eso queda para fase 2.
#
# Los umbrales de alerta (ej. % de margen mínimo) son configurables
# por el usuario en la UI, no están basados en ninguna cifra de
# documentos externos sobre el algoritmo de Meli.
# ============================================================

def _encontrar_fila_encabezado_ag(df_crudo, columnas_esperadas, max_filas=15):
    """
    Busca en las primeras filas cuál es la fila de encabezados real
    (los reportes de Meli traen texto descriptivo arriba del header).
    """
    for i in range(min(max_filas, len(df_crudo))):
        fila = df_crudo.iloc[i].astype(str).str.strip().tolist()
        coincidencias = sum(1 for c in columnas_esperadas if c in fila)
        if coincidencias >= 2:
            return i
    return None


def analizar_rentabilidad_ag(archivo_bytes, umbral_alerta_pct):
    """
    Lee detalle_ventas_de_mi_sistema (o similar): una línea por
    producto vendido, con precio, costo y % de profit ya calculado.
    """
    df = pd.read_excel(BytesIO(archivo_bytes), sheet_name=0, dtype=str)

    cols_necesarias = ["Codigo", "Articulo", "Cantidad", "Precio", "Precio Costo", "Profit %"]
    faltantes = [c for c in cols_necesarias if c not in df.columns]

    if faltantes:
        raise ValueError(
            "Este archivo no tiene la estructura esperada de detalle de ventas. "
            f"Faltan columnas: {', '.join(faltantes)}."
        )

    for col in ["Cantidad", "Precio", "Precio Costo", "Markup %", "Profit %", "Importe"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")

    alertas = df[df["Profit %"] < umbral_alerta_pct].copy()
    alertas = alertas.sort_values("Profit %")

    resumen = {
        "lineas_totales": len(df),
        "lineas_bajo_umbral": len(alertas),
        "lineas_a_perdida": int((df["Profit %"] < 0).sum()),
        "profit_promedio": round(df["Profit %"].mean(), 2) if df["Profit %"].notna().any() else None,
    }

    return df, alertas, resumen


def analizar_evolucion_negocio_ag(archivo_bytes):
    """
    Lee la hoja "Negocio" del Reporte de evolución del negocio de Meli:
    visitas, compradores, tasa de recompra, ventas por día.
    """
    xls = pd.ExcelFile(BytesIO(archivo_bytes))
    hoja_negocio = next((h for h in xls.sheet_names if "negocio" in h.lower()), xls.sheet_names[0])

    crudo = pd.read_excel(BytesIO(archivo_bytes), sheet_name=hoja_negocio, header=None, dtype=str)
    fila_header = _encontrar_fila_encabezado_ag(crudo, ["Fecha", "Visitas", "Compradores"])

    if fila_header is None:
        raise ValueError(
            "No se encontró la fila de encabezado (Fecha/Visitas/Compradores) "
            "en la hoja de Negocio de este archivo."
        )

    df = pd.read_excel(BytesIO(archivo_bytes), sheet_name=hoja_negocio, header=fila_header, dtype=str)
    df = df.dropna(subset=[df.columns[0]])

    for col in df.columns:
        if col != "Fecha":
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(",", "."), errors="coerce")

    df["Fecha"] = pd.to_datetime(df["Fecha"], format="%d/%m/%y", errors="coerce")
    df = df.sort_values("Fecha").reset_index(drop=True)

    return df


def analizar_top_productos_ag(archivo_bytes):
    """
    Lee el reporte "Ventas por Producto" de Meli: una fila por
    publicación, con unidades vendidas y monto facturado.
    """
    df = pd.read_excel(BytesIO(archivo_bytes), sheet_name=0, dtype=str)

    cols_necesarias = ["Productos", "Parent SKU", "Unidades Vendidas", "Monto"]
    faltantes = [c for c in cols_necesarias if c not in df.columns]

    if faltantes:
        raise ValueError(
            "Este archivo no tiene la estructura esperada de Ventas por Producto. "
            f"Faltan columnas: {', '.join(faltantes)}."
        )

    for col in ["Pedidos Concretados", "Unidades Vendidas", "Monto", "Ticket Promedio"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def mostrar_tablero_rentabilidad_ag(archivo):
    st.subheader("Rentabilidad por venta")

    umbral = st.slider(
        "Alertar ventas con margen por debajo de (%)",
        min_value=-50.0, max_value=50.0, value=15.0, step=1.0,
        key="ag_umbral_rentabilidad",
        help="Este número lo elegís vos; no es una regla de Mercado Libre."
    )

    try:
        with st.spinner("Analizando rentabilidad..."):
            df, alertas, resumen = analizar_rentabilidad_ag(archivo.getvalue(), umbral)
    except Exception as e:
        st.error("No se pudo leer el archivo.")
        st.exception(e)
        return

    c1, c2, c3 = st.columns(3)
    c1.metric("Líneas analizadas", f"{resumen['lineas_totales']:,}".replace(",", "."))
    c2.metric("Ventas a pérdida (margen < 0%)", resumen["lineas_a_perdida"])
    c3.metric(
        "Margen promedio",
        f"{resumen['profit_promedio']}%" if resumen["profit_promedio"] is not None else "N/D"
    )

    if not alertas.empty:
        st.warning(f"{len(alertas)} líneas con margen por debajo de {umbral}%.")

        columnas_mostrar = [c for c in ["Fecha", "Codigo", "Articulo", "Cantidad", "Precio", "Precio Costo", "Profit %"] if c in alertas.columns]

        st.dataframe(
            alertas[columnas_mostrar],
            use_container_width=True,
            hide_index=True,
            column_config={
                "Precio": st.column_config.NumberColumn(format="$ %.2f"),
                "Precio Costo": st.column_config.NumberColumn(format="$ %.2f"),
                "Profit %": st.column_config.NumberColumn(format="%.2f%%"),
            },
        )

        if "Articulo" in alertas.columns:
            peores = (
                alertas.groupby("Articulo")["Profit %"]
                .mean()
                .sort_values()
                .head(10)
            )
            st.caption("Artículos con peor margen promedio (top 10):")
            st.bar_chart(peores)

        st.divider()
        st.markdown("**Lecturas de estos datos**")

        if resumen["lineas_a_perdida"] > 0:
            st.write(
                f"- **{resumen['lineas_a_perdida']} líneas se vendieron a pérdida** "
                "(margen negativo). Revisá si es un error de precio cargado, un costo "
                "desactualizado, o una decisión consciente (ej. loss leader)."
            )
        if "Articulo" in alertas.columns and not alertas.empty:
            peor = alertas.iloc[0]
            st.write(
                f"- El peor caso puntual es **{peor['Articulo']}**, con margen de "
                f"{peor['Profit %']:.1f}%."
            )
    else:
        st.success(f"Ninguna línea por debajo del {umbral}% de margen.")


def mostrar_tablero_evolucion_ag(archivo):
    st.subheader("Evolución del negocio")

    try:
        with st.spinner("Leyendo evolución del negocio..."):
            df = analizar_evolucion_negocio_ag(archivo.getvalue())
    except Exception as e:
        st.error("No se pudo leer el archivo.")
        st.exception(e)
        return

    if df.empty:
        st.info("El archivo no tiene filas de datos.")
        return

    c1, c2, c3 = st.columns(3)
    c1.metric("Visitas totales del período", f"{int(df['Visitas'].sum()):,}".replace(",", "."))
    if "Cantidad de ventas" in df.columns:
        c2.metric("Ventas totales", f"{int(df['Cantidad de ventas'].sum()):,}".replace(",", "."))
    if "Tasa de recompra" in df.columns:
        c3.metric("Tasa de recompra promedio", f"{df['Tasa de recompra'].mean()*100:.1f}%")

    st.caption("Visitas por día:")
    st.line_chart(df.set_index("Fecha")["Visitas"])

    if "Cantidad de ventas" in df.columns:
        st.caption("Ventas por día:")
        st.line_chart(df.set_index("Fecha")["Cantidad de ventas"])

    if "Ventas brutas" in df.columns:
        conversion = (df["Cantidad de ventas"] / df["Visitas"] * 100).round(2)
        st.caption("Conversión estimada por día (ventas / visitas, %):")
        st.line_chart(conversion.rename("Conversión %"))

    with st.expander("Ver tabla completa"):
        st.dataframe(df, use_container_width=True, hide_index=True)

    st.divider()
    st.markdown("**Lecturas de estos datos**")

    dia_mas_visitas = df.loc[df["Visitas"].idxmax()]
    dia_menos_visitas = df.loc[df["Visitas"].idxmin()]
    st.write(
        f"- El día con más visitas fue **{dia_mas_visitas['Fecha'].strftime('%d/%m')}** "
        f"({int(dia_mas_visitas['Visitas'])} visitas); el de menos fue "
        f"**{dia_menos_visitas['Fecha'].strftime('%d/%m')}** ({int(dia_menos_visitas['Visitas'])})."
    )

    if "Cantidad de ventas" in df.columns:
        conversion_diaria = (df["Cantidad de ventas"] / df["Visitas"] * 100)
        mejor_idx = conversion_diaria.idxmax()
        peor_idx = conversion_diaria.idxmin()
        st.write(
            f"- Mejor conversión: **{df.loc[mejor_idx, 'Fecha'].strftime('%d/%m')}** "
            f"({conversion_diaria[mejor_idx]:.1f}%). Peor: "
            f"**{df.loc[peor_idx, 'Fecha'].strftime('%d/%m')}** ({conversion_diaria[peor_idx]:.1f}%). "
            "Si un día bajó fuerte sin caída de visitas, vale revisar stock, precio o quiebre "
            "de alguna publicación top ese día."
        )


def mostrar_tablero_top_productos_ag(archivo):
    st.subheader("Top productos")

    top_n = st.slider("Cuántos productos mostrar en cada ranking", 5, 30, 10, key="ag_top_n")

    try:
        with st.spinner("Leyendo ventas por producto..."):
            df = analizar_top_productos_ag(archivo.getvalue())
    except Exception as e:
        st.error("No se pudo leer el archivo.")
        st.exception(e)
        return

    c1, c2 = st.columns(2)
    c1.metric("Publicaciones con ventas", f"{len(df):,}".replace(",", "."))
    c2.metric("Monto total facturado", f"$ {df['Monto'].sum():,.0f}".replace(",", "."))

    tab1, tab2 = st.tabs(["Top por unidades vendidas", "Top por monto facturado"])

    columnas = [c for c in ["Productos", "Parent SKU", "Pedidos Concretados", "Unidades Vendidas", "Monto", "Ticket Promedio"] if c in df.columns]

    with tab1:
        top_unidades = df.sort_values("Unidades Vendidas", ascending=False).head(top_n)
        st.dataframe(
            top_unidades[columnas], use_container_width=True, hide_index=True,
            column_config={
                "Monto": st.column_config.NumberColumn(format="$ %.0f"),
                "Ticket Promedio": st.column_config.NumberColumn(format="$ %.2f"),
            },
        )
        st.bar_chart(top_unidades.set_index("Productos")["Unidades Vendidas"])

    with tab2:
        top_monto = df.sort_values("Monto", ascending=False).head(top_n)
        st.dataframe(
            top_monto[columnas], use_container_width=True, hide_index=True,
            column_config={
                "Monto": st.column_config.NumberColumn(format="$ %.0f"),
                "Ticket Promedio": st.column_config.NumberColumn(format="$ %.2f"),
            },
        )
        st.bar_chart(top_monto.set_index("Productos")["Monto"])

    st.divider()
    st.markdown("**Lecturas de estos datos**")

    total_monto = df["Monto"].sum()
    top10_monto = df.sort_values("Monto", ascending=False).head(10)["Monto"].sum()
    concentracion = (top10_monto / total_monto * 100) if total_monto else 0
    cola_larga = int((df["Unidades Vendidas"] <= 1).sum())
    pct_cola = (cola_larga / len(df) * 100) if len(df) else 0

    st.write(
        f"- Las 10 publicaciones que más facturan concentran el **{concentracion:.1f}%** "
        f"del monto total del período."
    )
    st.write(
        f"- **{cola_larga} de {len(df)} publicaciones ({pct_cola:.1f}%)** vendieron 0 o 1 "
        "unidad en este período. Vale revisar si conviene pausarlas, mejorarles el título/ficha "
        "técnica, o si son productos de rotación naturalmente baja (piezas raras, alto valor unitario)."
    )


# Firma de columnas de cada tipo de reporte que ya sabemos leer.
# Se busca la fila de encabezado que mejor matchee en cualquier hoja
# del archivo, sin que el usuario tenga que decir de antemano qué es.
FIRMAS_REPORTES_AG = {
    "rentabilidad": ["Codigo", "Articulo", "Cantidad", "Precio", "Precio Costo", "Profit %"],
    "top_productos": ["Productos", "Parent SKU", "Unidades Vendidas", "Monto"],
    "evolucion_negocio": ["Fecha", "Visitas", "Compradores"],
}


def _buscar_fila_encabezado_generica_ag(df_crudo, columnas_esperadas, max_filas=15):
    for i in range(min(max_filas, len(df_crudo))):
        fila = df_crudo.iloc[i].astype(str).str.strip().tolist()
        coincidencias = sum(1 for c in columnas_esperadas if any(c in str(v) for v in fila))
        if coincidencias >= 2:
            return i, coincidencias
    return None, 0


def detectar_tipo_reporte_ag(archivo_bytes):
    """
    Prueba las firmas conocidas contra cada hoja del archivo. Devuelve
    (tipo, nombre_hoja) del que mejor matcheó, o (None, None) si no
    reconoce ninguna estructura conocida — en ese caso no se debe
    forzar ningún análisis.
    """
    xls = pd.ExcelFile(BytesIO(archivo_bytes))

    mejor_tipo = None
    mejor_hoja = None
    mejor_score = 0

    for hoja in xls.sheet_names:
        crudo = pd.read_excel(BytesIO(archivo_bytes), sheet_name=hoja, header=None, dtype=str, nrows=15)

        for tipo, columnas in FIRMAS_REPORTES_AG.items():
            _, score = _buscar_fila_encabezado_generica_ag(crudo, columnas)
            if score > mejor_score:
                mejor_score = score
                mejor_tipo = tipo
                mejor_hoja = hoja

    if mejor_score >= 2:
        return mejor_tipo, mejor_hoja

    return None, None


def mostrar_herramienta_analista_general():
    st.title("Analista General")
    st.caption(
        "Subí el archivo y la app detecta sola qué tipo de reporte es. "
        "Por ahora reconoce: detalle de ventas (rentabilidad), evolución "
        "de negocio y ventas por producto, los 3 de Mercado Libre / tu sistema."
    )

    archivo = st.file_uploader(
        "Subí cualquiera de los reportes que uses",
        type=["xlsx"],
        key="ag_uploader_unico"
    )

    if archivo is None:
        return

    tipo, hoja = detectar_tipo_reporte_ag(archivo.getvalue())

    if tipo is None:
        st.error(
            "No reconozco la estructura de este archivo todavía. Reconozco: "
            "detalle de ventas con columnas Codigo/Articulo/Precio/Profit %, "
            "reporte de evolución de negocio (Fecha/Visitas/Compradores) y "
            "ventas por producto (Productos/Parent SKU/Unidades Vendidas/Monto). "
            "Si es un reporte distinto, decime cuál es para agregarle soporte."
        )
        return

    etiquetas = {
        "rentabilidad": "Rentabilidad por venta (tu sistema)",
        "top_productos": "Top productos (Meli)",
        "evolucion_negocio": "Evolución del negocio (Meli)",
    }
    st.success(f"Detectado como: **{etiquetas[tipo]}** (hoja: {hoja})")

    st.divider()

    if tipo == "rentabilidad":
        mostrar_tablero_rentabilidad_ag(archivo)
    elif tipo == "evolucion_negocio":
        mostrar_tablero_evolucion_ag(archivo)
    else:
        mostrar_tablero_top_productos_ag(archivo)


# ============================================================
# SIDEBAR / ROUTER PRINCIPAL
# ============================================================

st.sidebar.title("ML Toolkit")

herramienta = st.sidebar.selectbox(
    "Elegí herramienta",
    [
        "Actualizar Integraly",
        "Agente Publicador",
        "Best_price",
        "Analista General",
    ]
)

if herramienta == "Actualizar Integraly":
    mostrar_herramienta_integraly()

elif herramienta == "Agente Publicador":
    mostrar_herramienta_agente_publicador()

elif herramienta == "Best_price":
    mostrar_herramienta_best_price()

elif herramienta == "Analista General":
    mostrar_herramienta_analista_general()
